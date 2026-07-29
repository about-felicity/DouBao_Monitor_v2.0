import csv
import hashlib
import json
import os
import re
import shutil
import sqlite3
import sys
import tempfile
import time
from contextlib import contextmanager
from datetime import datetime, timezone, timedelta
from urllib.parse import urlparse
import urllib.request

import doubao_env_loader  # noqa: F401  loads API keys from local .env file
import doubao_question_aliases as qa
import doubao_brand_settings as brand_settings

# 固定使用中国时区 (UTC+8)，不受系统时区影响
CST = timezone(timedelta(hours=8))

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass


def now_str():
    return datetime.now(CST).isoformat(sep=" ", timespec="seconds")


def beijing_time_str(value="", *, fallback_now=False):
    """Normalize a timestamp to explicit UTC+8 without using emulator time."""
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value or "").strip()
        if not text:
            return now_str() if fallback_now else ""
        normalized = text.replace("Z", "+00:00").replace("z", "+00:00")
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            parsed = None
            for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
                try:
                    parsed = datetime.strptime(text[:19], pattern)
                    break
                except ValueError:
                    continue
            if parsed is None:
                return now_str() if fallback_now else text
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=CST)
    else:
        parsed = parsed.astimezone(CST)
    return parsed.isoformat(sep=" ", timespec="seconds")


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_CSV = os.path.join(BASE_DIR, "doubao_refs_result.csv")
OUT_XLSX = os.path.join(BASE_DIR, "doubao_refs_result.xlsx")
OUT_PRODUCTS_CSV = os.path.join(BASE_DIR, "doubao_products_result.csv")
OUT_ANSWERS_CSV = os.path.join(BASE_DIR, "doubao_answers_result.csv")
OUT_RUN_COUNTER = os.path.join(BASE_DIR, "doubao_run_counter.txt")
DATA_WRITE_LOCK = os.path.join(BASE_DIR, "doubao_product_data_write.lock")
ACCOUNT_TIME_FIELDS = [
    "account_uid",
    "account_uid_masked",
    "account_nickname",
    "web_account_uid",
    "source_device",
    "mumu_instance",
    "mumu_serial",
    "question_sent_at",
    "answer_completed_at",
    "captured_at",
    "source_uploaded_at",
    "receiver_received_at",
]
FIELDS = [
    "run_no",
    "run_time",
    "chat_id",
    "chat_title",
    "question",
    "page_url",
    *ACCOUNT_TIME_FIELDS,
    "status",
    "complete",
    "count",
    "expected_count",
    "index",
    "title",
    "href",
    "source",
    "extracted_at",
]
PRODUCT_FIELDS = [
    "run_no",
    "run_time",
    "chat_id",
    "chat_title",
    "question",
    "page_url",
    *ACCOUNT_TIME_FIELDS,
    "product_index",
    "product_name",
    "brand_name",
    "evidence",
    "product_count",
    "rank_type",
    "extraction_method",
    "review_status",
    "model",
    "reviewed_at",
    "answer_hash",
    "extracted_at",
]
ANSWER_FIELDS = [
    "run_no", "run_time", "chat_id", "chat_title", "question", "page_url",
    *ACCOUNT_TIME_FIELDS,
    "answer_text", "answer_hash", "review_status", "model", "reviewed_at", "extracted_at",
]
OLD_FIELDS = [
    "run_no",
    "run_time",
    "chat_id",
    "chat_title",
    "page_url",
    "status",
    "complete",
    "count",
    "expected_count",
    "index",
    "title",
    "href",
    "source",
    "extracted_at",
]


def u(*codes):
    return "".join(chr(code) for code in codes)


VIDEO = u(0x89c6, 0x9891)
ARTICLE = u(0x6587, 0x7ae0)
PRODUCT_PAGE = u(0x5546, 0x54c1, 0x9875)
OTHER = u(0x5176, 0x4ed6)
UNKNOWN = u(0x672a, 0x77e5)
DOUYIN = u(0x6296, 0x97f3)
DOUYIN_VIDEO = DOUYIN + VIDEO + u(0x94fe, 0x63a5)
MEDIA = u(0x5a92, 0x4f53)
NEWS_MEDIA = u(0x65b0, 0x95fb) + MEDIA
CHINA_NEWS = u(0x4e2d, 0x56fd, 0x65b0, 0x95fb, 0x7f51)
XN_DAILY = u(0x54b8, 0x5b81, 0x65e5, 0x62a5)
HEALTH_CHINA = u(0x5065, 0x5eb7, 0x4e2d, 0x56fd, 0x7f51)
TAOBAO = u(0x6dd8, 0x5b9d)
BY_TYPE = u(0x6309, 0x7c7b, 0x578b, 0x6c47, 0x603b)
BY_MEDIA = u(0x6309, 0x5a92, 0x4f53, 0x6c47, 0x603b)

OFFICIAL_HINTS = (
    ".gov.",
    ".gov/",
    ".edu.",
    ".edu/",
    ".org",
)
SOURCE_CACHE_JSON = os.path.join(BASE_DIR, "doubao_source_cache.json")
SOURCE_AI_CACHE_JSON = os.path.join(BASE_DIR, "doubao_source_ai_cache.json")
SOURCE_CONTENT_DB = os.path.join(BASE_DIR, "doubao_source_content.db")
DEBUG_LOG = os.path.join(BASE_DIR, "doubao_run_debug.log")
DEBUG_LOG_MAX_BYTES = 20 * 1024 * 1024
DEBUG_LOG_KEEP_BYTES = 2 * 1024 * 1024


def env_int(name, default):
    try:
        return max(1, int(os.environ.get(name, default)))
    except Exception:
        return default


def debug_log(message):
    if os.environ.get("DOUBAO_DEBUG_LOG", "1").strip() in ("0", "false", "FALSE", "no"):
        return
    try:
        if os.path.exists(DEBUG_LOG) and os.path.getsize(DEBUG_LOG) > DEBUG_LOG_MAX_BYTES:
            with open(DEBUG_LOG, "rb") as f:
                f.seek(max(0, os.path.getsize(DEBUG_LOG) - DEBUG_LOG_KEEP_BYTES))
                tail = f.read()
            with open(DEBUG_LOG, "wb") as f:
                f.write(b"[log rotated]\n")
                f.write(tail)
        with open(DEBUG_LOG, "a", encoding="utf-8") as f:
            f.write(now_str() + " " + str(message) + "\n")
    except Exception:
        pass


@contextmanager
def product_data_write_lock(timeout_seconds=20):
    """Serialize foreground capture and background product-review commits."""
    deadline = time.time() + timeout_seconds
    fd = None
    while fd is None:
        try:
            fd = os.open(DATA_WRITE_LOCK, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode("ascii", errors="ignore"))
        except FileExistsError:
            try:
                owner_pid = 0
                try:
                    with open(DATA_WRITE_LOCK, "r", encoding="ascii", errors="ignore") as lock_file:
                        owner_pid = int((lock_file.read() or "0").strip())
                except Exception:
                    owner_pid = 0
                owner_alive = process_is_running(owner_pid)
                # Large CSV schema/atomic rewrite operations can legitimately
                # take longer than 30 seconds. Never reap a lock owned by a
                # live process merely because it is old; that previously let
                # two captures allocate the same run_no. Only an abandoned
                # lock (dead owner), or a pid-less lock older than five
                # minutes, may be removed automatically.
                lock_age = time.time() - os.path.getmtime(DATA_WRITE_LOCK)
                if (owner_pid and not owner_alive) or (not owner_pid and lock_age > 5 * 60):
                    os.unlink(DATA_WRITE_LOCK)
                    continue
            except FileNotFoundError:
                continue
            if time.time() >= deadline:
                raise TimeoutError("product data write lock is busy")
            time.sleep(0.1)
    try:
        yield
    finally:
        try:
            os.close(fd)
        except Exception:
            pass
        # Windows virus scanners/readers can briefly keep the lock file open
        # after our descriptor is closed.  Retry the removal, but only while
        # the file still belongs to this process; never delete a successor's
        # freshly acquired lock.
        current_pid = str(os.getpid())
        for _ in range(20):
            try:
                with open(DATA_WRITE_LOCK, "r", encoding="ascii", errors="ignore") as lock_file:
                    if (lock_file.read() or "").strip() != current_pid:
                        break
                os.unlink(DATA_WRITE_LOCK)
                break
            except FileNotFoundError:
                break
            except PermissionError:
                time.sleep(0.1)


def process_is_running(pid):
    if not pid or pid == os.getpid():
        return bool(pid)
    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes

            process_query_limited_information = 0x1000
            still_active = 259
            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
            kernel32.OpenProcess.argtypes = [wintypes.DWORD, wintypes.BOOL, wintypes.DWORD]
            kernel32.OpenProcess.restype = wintypes.HANDLE
            kernel32.GetExitCodeProcess.argtypes = [wintypes.HANDLE, ctypes.POINTER(wintypes.DWORD)]
            kernel32.GetExitCodeProcess.restype = wintypes.BOOL
            kernel32.CloseHandle.argtypes = [wintypes.HANDLE]
            kernel32.CloseHandle.restype = wintypes.BOOL
            handle = kernel32.OpenProcess(
                process_query_limited_information, False, int(pid)
            )
            if not handle:
                return False
            try:
                exit_code = wintypes.DWORD()
                if not kernel32.GetExitCodeProcess(handle, ctypes.byref(exit_code)):
                    return False
                return exit_code.value == still_active
            finally:
                kernel32.CloseHandle(handle)
        except Exception:
            # A failed Windows process query must not make an abandoned lock
            # permanent.  Fall back to tasklist before treating it as alive.
            try:
                import subprocess
                output = subprocess.check_output(
                    ["tasklist", "/FI", "PID eq %s" % int(pid), "/FO", "CSV", "/NH"],
                    encoding="utf-8", errors="ignore",
                    creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                )
                return ('"%s"' % int(pid)) in output
            except Exception:
                return False
    try:
        os.kill(int(pid), 0)
        return True
    except OSError:
        return False


def load_payload(raw):
    raw = (raw or "").strip()
    if not raw:
      raise ValueError("empty json")
    return json.loads(raw)


def next_run_no():
    """Reserve and persist the next run number while the data lock is held.

    The counter must be written at allocation time. A capture with zero source
    links otherwise leaves no row in the refs CSV until its answer is archived,
    allowing a second capture to allocate the same number in that gap.
    """
    maximum = 0
    for path in (OUT_CSV, OUT_ANSWERS_CSV, OUT_PRODUCTS_CSV):
        if not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    try:
                        maximum = max(maximum, int(row.get("run_no") or 0))
                    except Exception:
                        pass
        except Exception:
            pass
    try:
        with open(OUT_RUN_COUNTER, "r", encoding="ascii") as f:
            maximum = max(maximum, int((f.read() or "0").strip()))
    except Exception:
        pass

    reserved = maximum + 1
    fd, temp_path = tempfile.mkstemp(prefix="doubao_run_counter_", suffix=".tmp", dir=BASE_DIR)
    os.close(fd)
    try:
        with open(temp_path, "w", encoding="ascii", newline="") as f:
            f.write(str(reserved))
            f.flush()
            os.fsync(f.fileno())
        os.replace(temp_path, OUT_RUN_COUNTER)
    finally:
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    return reserved


def chat_id_from_url(url):
    return (url or "").rstrip("/").split("/")[-1]


def normalize_bool(value):
    text = str(value or "").strip()
    if text.lower() in ("true", "1", "yes"):
        return "True"
    if text.lower() in ("false", "0", "no"):
        return "False"
    return text


PRODUCT_KEYWORDS = (
    "染发剂", "染发膏", "染发霜", "泡沫染", "染发乳", "染发露", "染眉", "眉毛增长液",
    "睫毛增长液", "睫毛精华", "生发液", "育发液", "洗发水", "精华", "面霜", "防晒",
)
PRODUCT_LINE_HINTS = (
    "推荐", "首选", "首迷", "综合", "优选", "平价", "高端", "入门", "产品", "品牌",
    "性价比", "敏感", "温和", "低敏", "新手", "院线", "成分",
)
PRODUCT_STOP_RE = re.compile(r"[。；;，,]\s*")


def strip_product_leading_noise(text):
    text = str(text or "").strip()
    text = re.sub(r"^[\s￥¥$]*\d+(?:\.\d+)?\s*/\s*\d+\s*(?:ml|mL|ML|g|G)\b", "", text).strip()
    text = re.sub(r"^\d+(?:\.\d+)?\s*(?:ml|mL|ML|g|G)\b", "", text).strip()
    text = re.sub(r"^[^\w\u4e00-\u9fff]+", "", text).strip()
    return text

RECOMMEND_QUESTION_HINTS = (
    "推荐", "怎么选", "如何选", "哪款", "哪个牌子", "什么牌子",
    "排行榜", "排行", "榜单", "清单", "合集", "对比", "选购",
)
NON_RECOMMEND_QUESTION_HINTS = (
    "怎么样", "评价", "评测", "安全吗", "安全么", "好用吗", "好不好",
    "成分安全吗", "成分安全", "靠谱吗", "是不是",
)


def normalize_payload_question(payload):
    raw_question = str(payload.get("question") or "").strip()
    normalized = qa.normalize_question_for_capture(
        raw_question,
        payload.get("chatTitle") or payload.get("title") or "",
    )
    if normalized and normalized != raw_question:
        debug_log("question normalized: " + repr(raw_question) + " -> " + repr(normalized))
        payload["question"] = normalized
    return normalized or raw_question


def is_recommendation_question(question):
    text = str(question or "").strip()
    if not text:
        return False
    if any(hint in text for hint in RECOMMEND_QUESTION_HINTS):
        return True
    if any(hint in text for hint in NON_RECOMMEND_QUESTION_HINTS):
        return False
    return False


def clean_product_name(text):
    text = str(text or "").strip()
    if not text:
        return ""
    text = re.sub(r"^[\s\-\*\u2022\d\.、\)\(]+", "", text)
    text = re.sub(r"^[^\w\u4e00-\u9fff]{1,4}", "", text)
    text = strip_product_leading_noise(text)
    text = re.sub(r"^(?:综合推荐|首选|推荐|低敏优选|入门性价比|成分党优选|院线级激活|产品|品牌)\s*[：:]\s*", "", text)
    if "|" in text or "｜" in text:
        parts = [p.strip() for p in re.split(r"[|｜]", text) if p.strip()]
        keyword_parts = [p for p in parts if any(k in p for k in PRODUCT_KEYWORDS)]
        if keyword_parts:
            text = keyword_parts[-1]
    if "：" in text or ":" in text:
        parts = [p.strip() for p in re.split(r"[：:]", text) if p.strip()]
        keyword_parts = [p for p in parts if any(k in p for k in PRODUCT_KEYWORDS)]
        if keyword_parts:
            text = keyword_parts[-1]
    text = PRODUCT_STOP_RE.split(text)[0].strip()
    text = re.sub(r"[（(]\s*(?:平价|高端|敏感|低敏|温和|均衡|全能|新手|孕妇|少刺激|家用)[^）)]*[）)]", "", text)
    text = re.sub(r"\s+", " ", text).strip(" -—：:，,。；;")
    if len(text) < 2 or len(text) > 40:
        return ""
    if not any(k in text for k in PRODUCT_KEYWORDS):
        return ""
    return text


def extract_products(answer_text):
    text = str(answer_text or "").replace("\r", "\n")
    if not text.strip():
        return []
    lines = []
    for raw in re.split(r"[\n\r]+|(?=\d+\s*[\.、])", text):
        raw = raw.strip()
        if not raw:
            continue
        # Some DOM text is flattened into long paragraphs. Split around likely
        # product heading markers without relying on a fixed brand list.
        chunks = re.split(r"(?=[✨🌿🛡🏆💡⚠]\s*)|(?=(?:综合推荐|首选|推荐|低敏优选|入门性价比|成分党优选|院线级激活)\s*[：:])", raw)
        lines.extend(chunk.strip() for chunk in chunks if chunk.strip())

    seen = set()
    products = []
    for line in lines:
        if not any(k in line for k in PRODUCT_KEYWORDS):
            continue
        short_product_heading = (
            len(line) <= 60
            and not re.search(r"[。；;，,：:]", line)
        )
        if not (
            any(h in line for h in PRODUCT_LINE_HINTS)
            or re.search(r"[|｜：:]", line)
            or short_product_heading
        ):
            continue
        name = clean_product_name(line)
        if not name or name in seen:
            continue
        seen.add(name)
        products.append({
            "product_name": name,
            "evidence": line[:240],
        })
    return products


def _ai_explicitly_disabled(name):
    value = os.environ.get(name, "").strip()
    return value in ("0", "false", "FALSE", "no")


def _has_api_key():
    return bool(os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("OPENAI_API_KEY"))


def should_use_ai_product_extractor(answer_text, products):
    # Default ON when an API key is configured, unless explicitly disabled.
    if not _has_api_key():
        return False
    if _ai_explicitly_disabled("DOUBAO_USE_AI_PRODUCT") or _ai_explicitly_disabled("DOUBAO_USE_AI_SOURCE"):
        return False

    text = str(answer_text or "").strip()
    if len(text) < 120:
        return False

    # Always call the model when the prose claims more products than the
    # rule parser found, so answers like "3 款精选推荐" are not truncated.
    claimed = declared_recommendation_count(text)
    if claimed and claimed > len(products or []):
        return True

    min_products = env_int("DOUBAO_AI_PRODUCT_MIN_PRODUCTS", 2)
    return (
        len(products or []) < min_products
        or has_suspicious_products(products)
        or os.environ.get("DOUBAO_USE_AI_PRODUCT_ALWAYS", "").strip() in ("1", "true", "TRUE", "yes")
    )


def is_suspicious_product_name(product_name):
    text = str(product_name or "").strip()
    if not text:
        return True
    suspicious_words = (
        "界面新闻", "给你挑了", "小提示", "小提醒", "小贴士", "提示", "贴士", "选购小贴士", "一句话", "注意", "提醒",
        "实测", "测评", "评测", "哪款", "哪个", "哪个好", "好闻", "好用",
        "清单", "指南", "排行榜", "榜单", "红黑榜", "科普", "区别",
        "价格", "图片", "品牌", "怎么样", "京东商城", "淘宝网", "网易网",
        "划重点", "先划重点", "合规", "≠", "靠多肽", "植萃滋养", "药物",
        "选睫毛增长液", "选眉毛增长液", "挑选睫毛增长液", "挑选眉毛增长液",
    )
    if any(word in text for word in suspicious_words):
        return True
    if re.search(r"^[^\w\u4e00-\u9fff]*\d+(?:\.\d+)?\s*/\s*\d+\s*(?:ml|mL|ML|g|G)\b", text):
        return True
    # A real product usually contains a product category word. If the extractor
    # produced an overly narrative phrase, ask AI to validate it instead.
    if not any(k in text for k in PRODUCT_NAME_KEYWORDS):
        return True
    return False


PRODUCT_ALIAS_RULES = (
    (("瑷尔博士", "益生菌", "面膜"), "瑷尔博士益生菌面膜"),
    (("珀莱雅", "源力", "面膜"), "珀莱雅源力面膜"),
    (("珀莱雅", "双抗", "面膜"), "珀莱雅双抗面膜"),
    (("珀莱雅", "水母", "面膜"), "珀莱雅水母面膜"),
    (("欧莱雅", "黑精华", "面膜"), "欧莱雅黑精华面膜"),
    (("理肤泉", "B5", "面膜"), "理肤泉 B5 面膜"),
    (("高姿", "蜂巢", "水库", "面膜"), "高姿蜂巢水库面膜"),
    (("润百颜", "白纱布", "面膜"), "润百颜白纱布面膜"),
    (("RNW", "玻尿酸", "面膜"), "RNW 玻尿酸面膜"),
    (("RNW", "四季", "发光", "面膜"), "RNW 四季发光面膜"),
    (("THESTARCHILD", "积雪草", "面膜"), "THESTARCHILD 积雪草面膜"),
    (("EIIO", "水光", "面膜"), "EIIO 水光面膜"),
    (("GIK", "PRP", "面膜"), "GIK PRP 胶原面膜"),
    (("金妮雅", "玻尿酸", "面膜"), "金妮雅玻尿酸面膜"),
    (("凯膜", "玻尿酸", "面膜"), "凯膜玻尿酸冻干粉面膜"),
    (("仁和匠心", "玻尿酸", "面膜"), "仁和匠心多重玻尿酸面膜"),
    (("澳贝妍", "377", "面膜"), "澳贝妍 377 美白面膜"),
    (("澳贝妍", "八杯水", "面膜"), "澳贝妍八杯水玻尿酸面膜"),
    (("八杯水", "玻尿酸", "面膜"), "八杯水玻尿酸面膜"),
    (("韩方五谷", "377", "面膜"), "韩方五谷 377 美白面膜"),
    (("美美的天空", "积雪草", "面膜"), "美美的天空积雪草面膜"),
    (("薇诺娜", "屏障", "面膜"), "薇诺娜屏障修护面膜"),
    (("可复美", "胶原蛋白", "修护贴"), "可复美重组胶原蛋白修护贴"),
    (("玻尿酸", "冻干粉", "面膜"), "凯膜玻尿酸冻干粉面膜"),
    (("神秘博士", "二裂酵母", "面膜"), "神秘博士二裂酵母面膜"),
    (("LiLiA", "玻尿酸", "面膜"), "LiLiA 8D 玻尿酸面膜"),
)

# Known master brands used to split compact Chinese names like
# "施华蔻怡然染发霜" into "施华蔻 怡然染发霜" and to infer the correct brand.
KNOWN_BRAND_NAMES = (
    # 染发
    "施华蔻", "欧莱雅", "花王", "梵玢", "FBCY", "陶博士", "章华", "美源",
    "科熙本", "忆丝芸", "青羽雀",
    # 睫毛 / 眉毛
    "修正", "朵妆", "AVANCE", "Cavilla", "卡维拉", "GeraX", "REVITALASH",
    "RevitaLash", "悦密佳", "斯必申", "Almea", "蓝鲸眼泪", "UKISS", "MAVALA",
    "Mavala",
    # 面膜 / 其他
    "瑷尔博士", "珀莱雅", "理肤泉", "高姿", "润百颜", "RNW", "THESTARCHILD",
    "EIIO", "GIK", "金妮雅", "凯膜", "仁和匠心", "澳贝妍", "韩方五谷",
    "美美的天空", "神秘博士", "LiLiA", "薇诺娜", "可复美",
)


def normalize_known_product_alias(product_name):
    text = str(product_name or "").strip()
    # Normalize repeated/model-expanded brand prefixes such as
    # "梵玢 FBCY FBCY 梵玢眉毛精华液" to one clean master-brand prefix.
    if re.match(r"^(?:梵正|梵玢|FBCY)", text, re.I):
        remainder = re.sub(r"^(?:(?:梵正|梵玢|FBCY)\s*)+", "", text, flags=re.I)
        text = ("梵玢 FBCY " + remainder).strip()
    text = re.sub(r"^Mavala", "MAVALA", text, flags=re.I)
    text = re.sub(r"^ALMEA", "Almea", text, flags=re.I)
    text = re.sub(r"^道和时尚", "道和", text)
    # Insert a space after a known brand prefix for compact names like
    # "施华蔻怡然染发霜" -> "施华蔻 怡然染发霜", so product and brand
    # aggregation stays consistent across runs.
    for brand in sorted(KNOWN_BRAND_NAMES, key=len, reverse=True):
        if re.search(r"^" + re.escape(brand) + r"\s*", text, re.IGNORECASE):
            remainder = re.sub(r"^" + re.escape(brand) + r"\s*", "", text, flags=re.IGNORECASE)
            text = (brand + " " + remainder).strip()
            break
    # Ensure a space between a leading Chinese brand and the rest of the name
    # so "朵妆多肽睫毛滋养液" and "朵妆 多肽睫毛滋养液" aggregate together.
    text = re.sub(r"^(朵妆|悦密佳|卡维拉|韩方五谷|澳贝妍|金妮雅|凯膜|仁和匠心|LiLiA|RNW|高姿|润百颜|美美的天空|神秘博士)([\u4e00-\u9fffA-Za-z])", r"\1 \2", text)
    # Normalize "精华" -> "精华液" for eyebrow/lash products so variants merge.
    # Handle both "睫毛精华液" and "睫毛臻萃精华" style names.
    if "精华液" not in text and any(kw in text for kw in ("眉毛", "睫毛")):
        text = re.sub(r"精华\b", "精华液", text)
    # Normalize Cavilla / 卡薇拉 / 卡维拉 variants to the canonical Chinese brand.
    text = re.sub(r"(?i)(?:CAVILLA|Cavilla)\s*[卡咖]薇拉?", "卡维拉", text)
    text = re.sub(r"(?i)[卡咖]薇拉\s*(?:CAVILLA|Cavilla)", "卡维拉", text)
    text = re.sub(r"卡薇拉", "卡维拉", text)
    # Normalize common casing variants.
    text = re.sub(r"(?i)\bgerax\b", "GeraX", text)
    # Normalize minoxidil presentation order.
    text = re.sub(r"^(斯必申)\s+米诺地尔搽剂\s+(\d+%)", r"\1 \2 米诺地尔搽剂", text)
    for tokens, canonical in PRODUCT_ALIAS_RULES:
        if all(token.lower() in text.lower() for token in tokens):
            return canonical
    return re.sub(r"\s+", " ", text).strip()


def infer_product_name_from_payload_context(product_name, evidence, payload):
    """Use same-round references to fill brands omitted by the answer text."""
    name = normalize_known_product_alias(product_name)
    text = f"{name} {evidence or ''}"
    refs_text = " ".join(
        str(item.get("title") or "") for item in (payload.get("items") or []) if isinstance(item, dict)
    )

    # The answer often says only "玻尿酸洗发水（紫瓶）", while the same-round
    # reference title says it is 欧莱雅. Prefixing only under this context avoids
    # treating every generic "玻尿酸洗发水" as 欧莱雅.
    if "玻尿酸洗发水" in text and "欧莱雅" in refs_text and "欧莱雅" not in name:
        return normalize_known_product_alias("欧莱雅 " + name)

    return name


def has_suspicious_products(products):
    return any(is_suspicious_product_name(item.get("product_name", "")) for item in (products or []))


def canonical_ai_brand(brand, product_name="", sub_brand=""):
    """Normalize model brand hierarchy to the dashboard aggregation level."""
    brand = re.sub(r"\s+", " ", str(brand or "")).strip()
    combined = " ".join((brand, str(product_name or ""), str(sub_brand or "")))
    if "花王莉婕" in combined or brand.lower() == "liese" or brand == "莉婕":
        return "花王"
    if brand.startswith("梵玢"):
        return "梵玢 FBCY"
    if brand.casefold() == "mavala":
        return "MAVALA"
    if brand.casefold() == "almea":
        return "Almea"
    if "乌斯玛" in brand:
        return "新疆乌斯玛"
    if brand.startswith("道和"):
        return "道和"
    if brand.casefold() == "ryo" or brand in ("吕", "紫吕", "吕（Ryo）", "吕RYO") or re.search(r"(?:紫吕|\bRYO\b)", combined, re.I):
        return "吕RYO"
    if brand in ("康王", "拜耳康王", "拜耳") and "康王" in combined:
        return "康王"
    try:
        import doubao_dashboard_server as dashboard
        return dashboard.canonical_brand_name(brand)
    except Exception:
        return brand


def normalize_ai_product_brand_prefix(product_name, canonical_brand):
    """Collapse repeated Chinese/English brand aliases to one master prefix."""
    product = str(product_name or "").strip()
    brand = str(canonical_brand or "").strip()
    if not product or not brand:
        return product
    aliases = {brand}
    try:
        import doubao_dashboard_server as dashboard
        aliases.update(dashboard.aliases_for_brand(brand))
    except Exception:
        pass
    remainder = product
    for _ in range(4):
        changed = False
        for alias in sorted(aliases, key=len, reverse=True):
            match = re.match(
                r"^\s*" + re.escape(alias) + r"(?=$|[\s·/（）()\-—+])",
                remainder,
                re.I,
            )
            if not match:
                continue
            remainder = remainder[match.end():].lstrip(" \t·/（）()-—+")
            changed = True
            break
        if not changed:
            break
    return (brand + (" " + remainder if remainder else "")).strip()


def normalize_ai_products(parsed):
    if not isinstance(parsed, dict):
        return []
    raw_items = parsed.get("products") or parsed.get("items") or []
    if not isinstance(raw_items, list):
        return []
    products = []
    seen = set()
    for position, raw in enumerate(raw_items, 1):
        if not isinstance(raw, dict):
            continue
        product = str(raw.get("product_name") or raw.get("product") or raw.get("name") or "").strip()
        brand = str(raw.get("brand") or raw.get("parent_brand") or "").strip()
        sub_brand = str(raw.get("sub_brand") or "").strip()
        brand = canonical_ai_brand(brand, product, sub_brand)
        if brand and product:
            product = normalize_ai_product_brand_prefix(product, brand)
        product = re.sub(r"\s+", " ", product).strip(" -—：:；;。")
        product = strip_product_leading_noise(product)
        product = normalize_known_product_alias(product)
        # AI extraction is deliberately less restrictive than the legacy rule
        # parser.  Product nicknames such as "小黑瓶" do not necessarily contain
        # a category keyword, but are valid when the model can point to answer
        # evidence.  Still reject narrative fragments and unsafe-long strings.
        if not product or len(product) > 80:
            continue
        # “防晒”既是品类词，也是真实防晒商品名的一部分。旧逻辑按子串
        # 禁止它，导致“珀莱雅盾护防晒”“珂润润浸保湿防晒”等被模型
        # 正确抽取后又在这里丢弃。AI 已提供原文证据，因此只拒绝纯品类名。
        ai_forbidden_hints = (hint for hint in PRODUCT_FORBIDDEN_NAME_HINTS if hint != "防晒")
        if any(hint in product for hint in ai_forbidden_hints):
            continue
        if re.fullmatch(r"防晒(?:霜|乳|液|产品)?", product):
            continue
        evidence = str(raw.get("evidence") or raw.get("reason") or raw.get("line") or product).strip()
        if not evidence or len(evidence) < 2:
            continue
        try:
            rank = int(raw.get("rank") or position)
        except Exception:
            rank = position
        if rank < 1:
            rank = position
        rank_type = str(raw.get("rank_type") or "appearance_order").strip().lower()
        if rank_type not in ("explicit_rank", "appearance_order"):
            rank_type = "appearance_order"
        # Treat spacing/punctuation-only variants as the same recommendation,
        # e.g. "质润 二硫化硒洗发水" and "质润二硫化硒洗发水".
        key = (
            re.sub(r"[\s\-_—·+]+", "", brand).casefold(),
            re.sub(r"[\s\-_—·+]+", "", product).casefold(),
        )
        if key in seen:
            continue
        seen.add(key)
        products.append({
            "product_name": product,
            "brand_name": brand,
            "evidence": evidence[:240],
            "rank": rank,
            "rank_type": rank_type,
        })
    return products


def declared_recommendation_count(answer_text):
    """Return an explicit list size such as “给你挑了 3 款”, if present."""
    text = str(answer_text or "")
    patterns = (
        r"(?:给你|为你|帮你)(?:挑|选|推荐|整理)?了?\s*(\d{1,2})\s*款",
        r"(?:这里|下面)(?:给你)?(?:挑|选|推荐|整理)了?\s*(\d{1,2})\s*款",
        r"(?:精选|推荐|整理|盘点|测评|亲测|实测)\s*(\d{1,2})\s*款",
        r"(?:分|按).{0,6}(?:类|档|场景|需求|肤质|预算).{0,10}(\d{1,2})\s*款",
        r"共\s*(\d{1,2})\s*款",
        r"(\d{1,2})\s*款(?:精选|推荐|安利|闭眼入|必入|清单|合集|总结|测评|横评)",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            count = int(match.group(1))
            if 1 < count <= 20:
                return count
    return 0


def ensure_complete_ai_products(answer_text, parsed, products):
    """Ensure normalization did not discard anything the model extracted.

    The answer can be internally inconsistent (for example, say “3款” but
    actually name only two).  Therefore the prose claim is diagnostic only;
    the hard guarantee is that every raw model item survives normalization.
    """
    raw_items = []
    if isinstance(parsed, dict):
        raw_items = parsed.get("products") or parsed.get("items") or []
    raw_count = sum(
        1 for item in raw_items
        if isinstance(item, dict)
        and str(item.get("product_name") or item.get("product") or item.get("name") or "").strip()
    )
    if raw_count != len(products or []):
        raise ValueError(
            "normalization dropped model products: raw=%d normalized=%d"
            % (raw_count, len(products or []))
        )
    claimed = declared_recommendation_count(answer_text)
    if claimed and claimed != len(products or []):
        debug_log(
            "answer claimed %d products but explicitly named/model-extracted %d; "
            "accepting the explicit names" % (claimed, len(products or []))
        )
    return products


def strip_reference_prefix(text):
    """Remove reference/noise blocks from the captured answer.

    Drops both the leading reference-title block Doubao prepends and trailing
    sections like '相关视频' / '相关推荐' that contain video/article titles,
    hashtags, and other unrelated content. These blocks are not the assistant's
    product recommendations and confuse the extractor, so we keep only the
    prose recommendation body.
    """
    text = str(text or "").strip()
    if not text:
        return text

    # 1) Drop the initial quoted keyword section: "关键词1"、"关键词2"
    text = re.sub(r'^["""][^"""]+["""](?:\s*、\s*["""][^"""]+["""])*\s*', '', text)

    # 2) Drop consecutive reference-title lines that end with hashtags.
    #    Stop as soon as we hit a prose line that does not end with #tags.
    lines = text.splitlines()
    cleaned_lines = []
    skipping = True
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        # A reference title line ends with one or more #hashtag chunks.
        if skipping and re.search(r'(?:#\S+\s*)+$', stripped):
            continue
        skipping = False
        cleaned_lines.append(line)

    # 3) Drop trailing related-content sections and everything after them.
    #    These sections list video/article titles and are not part of the
    #    assistant's own product recommendations.
    TRAILING_SECTION_HEADERS = (
        "相关视频", "相关推荐", "相关文章", "相关笔记", "相关问答",
        "相关资讯", "相关搜索", "相关内容", "参考资料", "参考链接",
        "延伸阅读", "相关视频推荐",
    )
    cutoff = len(cleaned_lines)
    for i, line in enumerate(cleaned_lines):
        stripped = line.strip()
        if any(stripped.startswith(marker) for marker in TRAILING_SECTION_HEADERS):
            cutoff = i
            break
    cleaned_lines = cleaned_lines[:cutoff]

    # 4) If the entire text was reference titles, fall back to the original.
    result = "\n".join(cleaned_lines).strip()
    if not result:
        return text
    return result


def build_product_prompt(answer_text):
    cleaned = strip_reference_prefix(answer_text)
    return {
        "task": "Extract products explicitly recommended by an AI answer.",
        "rules": [
            "Return strict JSON only.",
            "Read the complete answer body, including all paragraphs and list items. Extract every real product explicitly recommended by the assistant.",
            "A product may be named by a nickname or a full name; preserve the most specific brand + product name present in the body. Never invent or expand a name using outside knowledge.",
            "Do not extract reference titles, web page titles, article/video titles, category words, warnings, feature bullets, price/image/brand search titles, or general advice.",
            "Keep the rank from the answer. Use rank_type=explicit_rank only when the body gives a numeric/ordinal rank; otherwise use rank_type=appearance_order and rank by first recommendation appearance.",
            "For dashboard aggregation, brand must be the stable master brand, not a concatenated master-brand + sub-brand/product-line name. Keep the sub-brand or series in product_name and sub_brand. Example: 花王莉婕泡沫染发剂 => brand=花王, sub_brand=莉婕, product_name=花王莉婕泡沫染发剂.",
            "If the answer only gives a sub-brand, you may use stable brand ownership knowledge to normalize brand, but never invent or expand product_name beyond the answer text.",
            "Within one answer, the same brand must always use exactly one normalized brand string. Merge casing, spacing, decorative-symbol, Chinese/English co-name, and common typo variants instead of returning separate brands.",
            "Prefer the stable Chinese consumer-facing master brand when Chinese and English names refer to the same brand. Examples: Fresh/馥蕾诗=>馥蕾诗, Freiol/福来=>福来, Moroccanoil/摩洛哥油=>摩洛哥油, Cavilla/CAVILLA/卡薇拉=>卡维拉.",
            "Normalize known formatting variants: DS 实验室=>DS实验室, Spēs=>Spes, vsve=>VSVE, +OKSS+/OKSS+=>OKSS. Decorative plus signs are not part of the master brand.",
            "Do not use a product line or duplicated prefix as brand. Examples: 仁和匠心=>仁和, 章华汉草=>章华, 甘椰植萃=>甘椰, 因士柔酸=>因士.",
            "Do not output the same recommendation twice under brand aliases. Deduplicate by normalized brand + normalized product identity.",
            "If the answer only evaluates one named product and does not recommend a list, return an empty products array.",
            "Each item must have rank, brand, sub_brand, product_name, evidence.",
        ],
        "output_schema": {
            "products": [
                {
                    "rank": 1,
                    "rank_type": "explicit_rank or appearance_order",
                    "brand": "用于面板聚合的主品牌名",
                    "sub_brand": "子品牌或产品系列名；没有则为空字符串",
                    "product_name": "品牌 + 产品名",
                    "evidence": "原文中证明它被推荐的短句",
                }
            ]
        },
        "answer_text": cleaned[:6000],
    }


def model_response_text(data):
    """Accept both Anthropic Messages and OpenAI-compatible response shapes."""
    content = ""
    for part in data.get("content") or []:
        if isinstance(part, dict) and part.get("type") == "text":
            content += part.get("text", "")
    if content:
        return content
    choices = data.get("choices") or []
    if choices and isinstance(choices[0], dict):
        message = choices[0].get("message") or {}
        value = message.get("content") or choices[0].get("text") or ""
        if isinstance(value, list):
            return "".join(str(item.get("text") or "") for item in value if isinstance(item, dict))
        return str(value)
    return ""


def call_anthropic_product_extractor(answer_text):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return None
    base_url = os.environ.get("ANTHROPIC_BASE_URL", "https://api.deepseek.com/anthropic").rstrip("/")
    model = os.environ.get("DOUBAO_PRODUCT_AI_MODEL", "deepseek-v4-flash").strip() or "deepseek-v4-flash"
    attempts = env_int("DOUBAO_PRODUCT_AI_ATTEMPTS", 1)
    for attempt in range(1, attempts + 1):
        body = {
            "model": model,
            # Keep enough room for every recommendation and its evidence.
            "max_tokens": 3000,
            "temperature": 0,
            "system": "Return one valid JSON object only. Do not use markdown or commentary.",
            "messages": [
                {
                    "role": "user",
                    "content": json.dumps(build_product_prompt(answer_text), ensure_ascii=False),
                }
            ],
        }
        try:
            req = urllib.request.Request(
                base_url + "/v1/messages",
                data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "Content-Type": "application/json",
                },
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=env_int("DOUBAO_AI_PRODUCT_TIMEOUT", 45)) as resp:
                data = json.loads(resp.read().decode("utf-8", errors="ignore"))
            content = model_response_text(data)
            if not content.strip():
                raise ValueError("empty model response")
            parsed = parse_product_json(content)
            products = normalize_ai_products(parsed)
            return ensure_complete_ai_products(answer_text, parsed, products)
        except Exception as exc:
            debug_log(
                "AI product extractor attempt %d/%d failed: %r"
                % (attempt, attempts, exc)
            )
            if attempt < attempts:
                time.sleep(min(2 * attempt, 4))
    return None


def call_openai_product_extractor(answer_text):
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return None
    model = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": "Extract recommended products. Output strict JSON only."},
            {"role": "user", "content": json.dumps(build_product_prompt(answer_text), ensure_ascii=False)},
        ],
        "temperature": 0,
    }
    try:
        base_url = os.environ.get("OPENAI_BASE_URL", "https://api.openai.com").rstrip("/")
        endpoint = base_url + ("/chat/completions" if base_url.endswith("/v1") else "/v1/chat/completions")
        req = urllib.request.Request(
            endpoint,
            data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
            headers={
                "Authorization": "Bearer " + api_key,
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=env_int("DOUBAO_AI_PRODUCT_TIMEOUT", 45)) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="ignore"))
        parsed = parse_json_object(data["choices"][0]["message"]["content"])
        products = normalize_ai_products(parsed)
        return ensure_complete_ai_products(answer_text, parsed, products)
    except Exception as exc:
        debug_log("OpenAI product extractor failed: " + repr(exc))
        return None


def product_ai_mode():
    """Use verified AI when configured; stay useful offline with a labeled rule fallback."""
    configured = os.environ.get("DOUBAO_PRODUCT_AI_MODE", "").strip().lower()
    if configured:
        return configured if configured in ("required", "fallback", "off") else "required"
    has_model_key = bool(
        os.environ.get("ANTHROPIC_API_KEY")
        or os.environ.get("OPENAI_API_KEY")
    )
    mode = "required" if has_model_key else "fallback"
    return mode if mode in ("required", "fallback", "off") else "required"


def product_ai_model_label():
    if os.environ.get("ANTHROPIC_API_KEY"):
        return os.environ.get("DOUBAO_PRODUCT_AI_MODEL", "deepseek-v4-flash").strip() or "deepseek-v4-flash"
    if os.environ.get("OPENAI_API_KEY"):
        return os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
    return ""


def review_products_with_ai(answer_text):
    """Return (products, status, method, model).

    A successful empty list is an AI-reviewed result: it means the answer did
    not recommend a product.  A failed call is kept distinct so it can be
    retried instead of silently falling back to inaccurate regex output.

    The vocabulary/rule result is only a failure fallback.  A partial
    vocabulary hit must never skip the model because that silently drops
    unknown brands from otherwise valid multi-product answers.
    """
    text = strip_reference_prefix(answer_text)
    rule_products = extract_products(text)
    historical_products = match_historical_brands(text)
    mode = product_ai_mode()
    if not text:
        return [], "no_answer", "none", ""
    if mode == "off":
        products = historical_products if historical_products else rule_products
        return products, "rule_unverified", "rule", ""

    debug_log(
        "AI product review start; answer_len=" + str(len(text))
        + " rule_count=" + str(len(rule_products))
        + " historical_count=" + str(len(historical_products))
    )
    result = call_anthropic_product_extractor(text)
    method = "anthropic"
    if result is None:
        result = call_openai_product_extractor(text)
        method = "openai"
    if result is not None:
        # A model response can contain a valid product while omitting its
        # brand. Fill only names present in the curated canonical vocabulary;
        # never promote an arbitrary product name into a brand.
        try:
            import doubao_dashboard_server as dashboard
            for item in result:
                if str(item.get("brand_name") or "").strip():
                    continue
                product_name = str(item.get("product_name") or "").strip()
                folded_product = product_name.casefold()
                canonical = ""
                for aliases, brand in sorted(
                    dashboard.BRAND_ALIAS_RULES,
                    key=lambda rule: max(len(alias) for alias in rule[0]),
                    reverse=True,
                ):
                    if any(folded_product.startswith(alias.casefold()) for alias in aliases):
                        canonical = brand
                        break
                if not canonical:
                    inferred = infer_brand_from_product_name(product_name)
                    canonical = dashboard.canonical_brand_name(inferred)
                if canonical in dashboard.KNOWN_BRANDS:
                    item["brand_name"] = canonical
        except Exception:
            pass
        debug_log("AI product review completed; ai_count=" + str(len(result)))
        return result, "ai_verified", method, product_ai_model_label()

    if mode == "fallback":
        debug_log("AI product review failed; using explicitly allowed rule fallback")
        fallback_products = historical_products if historical_products else rule_products
        return fallback_products, "rule_unverified", "rule_fallback", ""
    debug_log("AI product review failed; queued as pending and omitted from product stats")
    return [], "ai_pending", "pending", product_ai_model_label()


def extract_products_with_ai_fallback(answer_text):
    """Compatibility wrapper for callers outside this module."""
    return review_products_with_ai(answer_text)[0]


PRODUCT_NAME_KEYWORDS = (
    "染发剂", "染发膏", "染发霜", "泡沫染", "染发乳", "染发露",
    "沐浴油", "沐浴精油", "沐浴露", "身体油", "护理油",
    "眉毛增长液", "眉毛精华液", "眉毛滋养液", "睫毛增长液", "睫毛精华液", "睫毛滋养液", "睫毛精华",
    "生发液", "育发液", "洗发水", "精华液", "精华", "滋养液", "面膜", "面霜", "防晒霜", "防晒乳",
    "米诺地尔", "搽剂",
)
PRODUCT_DETAIL_MARKERS = (
    " 特点", "特点：", "特点:", " 核心", "核心：", "核心:",
    " 适合", "适合：", "适合:", " 成分", "成分：", "成分:",
    " 功效", "功效：", "功效:", " 规格", "规格：", "规格:",
    " 价格", "价格：", "价格:", " 参考价", "参考价：", "参考价:",
    " 肤感", "肤感：", "肤感:", " 用法", "用法：", "用法:",
    " 推荐理由", "推荐理由：", "推荐理由:",
)
PRODUCT_HEADING_HINTS = (
    "推荐", "首选", "优选", "平价", "高端", "入门", "性价比", "新手", "敏感",
    "干皮", "油皮", "敏感肌", "修护", "香氛", "日常", "经典", "综合",
)
PRODUCT_FORBIDDEN_NAME_HINTS = (
    "#", "测评", "指南", "红黑榜", "榜单", "排行榜", "选购", "怎么选", "推荐】",
    "推荐”", "推荐\"", "参考", "网页", "热门沐浴油", "一句话点评",
    "不可直接", "可直接用", "不是产品", "区别", "科普", "几个", "提升", "幸福感",
    "评测", "精选评测", "价格", "图片", "品牌", "怎么样", "京东商城", "淘宝网",
    "商品网", "网易网", "手机网易网", "重要提醒", "必须", "否则", "极易", "反黑",
    "防晒", "先防晒", "效果大打折扣", "正品", "国货", "产品晚霜", "补水保湿",
    "实测", "测出来", "哪个好", "哪款好", "哪个味道", "好闻吗", "好用吗", "热门",
    "小提示", "小提醒", "小贴士", "每周", "每次",
    "需坚持", "才见效", "有泛红", "有刺激", "有干燥", "有风险", "属药品", "是药品",
    "给你挑了", "给你 3 款", "推荐 3 款", "分 3 款", "按需求", "按新手",
)

PRODUCT_NOISE_PATTERNS = (
    r"^[^，。；;]{0,8}(?:评测|测评|指南|排行榜|榜单|选购|科普)",
    r"(?:价格|图片|品牌|怎么样).{0,10}(?:京东|淘宝|商城|商品网|网易网)",
    r"(?:京东商城|淘宝网|手机网易网|网易网|商品网)$",
    r"^(?:重要提醒|注意|提醒|小提示|小提醒|小贴士)",
    r"(?:面膜每周|面膜一周|每周\s*2-3\s*次|每次\s*15)",
    r"(?:必须|否则|极易|反黑|先防晒|防晒).*(?:面霜|防晒霜)",
    r"(?:实测|测评|评测|哪个好|哪款好|哪个味道|好闻吗|好用吗)",
)


def _line_after_leading_markers(line):
    """Return line with leading list markers / emojis stripped.

    Skip filters need to match the semantic start of a line, but Doubao
    answers often prefix headings with emoji or numbers. Stripping those
    before testing makes the filters reliable.
    """
    text = str(line or "").strip()
    text = re.sub(r"^[✨🌿🛡🏆💡⚠✅⭐💧🔥💰❗🌟🔴🟡🟢🔵👍💪🧴]\s*", "", text)
    text = re.sub(r"^[^\n]{0,4}(?:\d+|[一二三四五六七八九十])[\.\、]\s*", "", text)
    text = re.sub(r"^\([\d一二三四五六七八九十]+\)\s*", "", text)
    text = re.sub(r"^[\d一二三四五六七八九十]+[)\.\、]\s*", "", text)
    return text.strip()


def split_answer_product_lines(answer_text):
    text = str(answer_text or "").replace("\r", "\n")
    if not text.strip():
        return []
    text = re.sub(r"(?<!\n)(?=(?:\d+|[一二三四五六七八九十])[\.\、]\s*)", "\n", text)
    text = re.sub(r"(?<!\n)(?=[✨🌿🛡🏆💡⚠✅⭐💧🔥💰]\s*)", "\n", text)
    text = re.sub(r"(?<!\n)(?=(?:综合推荐|首选|推荐|低敏优选|入门性价比|成分党优选|院线级激活|经典香氛|日常保湿|干皮|油皮|敏感肌)\s*[：:])", "\n", text)
    return [line.strip() for line in re.split(r"[\n\r]+", text) if line.strip()]


def cut_product_detail(text):
    text = str(text or "").strip()
    positions = [text.find(marker) for marker in PRODUCT_DETAIL_MARKERS if marker in text]
    positions = [pos for pos in positions if pos > 0]
    if positions:
        text = text[:min(positions)].strip()
    # Stop at the first sentence-ending punctuation or comma; commas in
    # recommendation headings are usually separators, not part of the name.
    text = re.split(r"[。；;，,]\s*", text)[0].strip()
    return text


def clean_recommended_product_name(text):
    text = str(text or "").strip()
    if not text:
        return ""
    # Strip ASCII list markers only; Chinese number words like "三橡树" are
    # part of brand names and must not be removed.  Ordinal markers such as
    # "一、" are already stripped by _line_after_leading_markers.
    text = re.sub(r"^[\s\-\*\u2022\d\.、\)\(]+", "", text)
    text = re.sub(r"^[^\w\u4e00-\u9fff]{1,4}", "", text)
    text = strip_product_leading_noise(text)
    text = cut_product_detail(text)
    if any(hint in text for hint in PRODUCT_FORBIDDEN_NAME_HINTS):
        return ""
    if any(re.search(pattern, text) for pattern in PRODUCT_NOISE_PATTERNS):
        return ""
    if "｜" in text or "|" in text:
        parts = [p.strip() for p in re.split(r"[|｜]", text) if p.strip()]
        keyword_parts = [p for p in parts if any(k in p for k in PRODUCT_NAME_KEYWORDS)]
        text = keyword_parts[-1] if keyword_parts else parts[-1]
    if "：" in text or ":" in text:
        parts = [p.strip() for p in re.split(r"[：:]", text) if p.strip()]
        keyword_parts = [p for p in parts if any(k in p for k in PRODUCT_NAME_KEYWORDS)]
        text = keyword_parts[-1] if keyword_parts else parts[-1]
    text = cut_product_detail(text)
    if any(hint in text for hint in PRODUCT_FORBIDDEN_NAME_HINTS):
        return ""
    if any(re.search(pattern, text) for pattern in PRODUCT_NOISE_PATTERNS):
        return ""
    # Drop descriptive parenthetical tags that Doubao appends to recommendations
    # (e.g. "朵妆 多肽睫毛滋养液（高性价比）"). Keep the product name clean for
    # dashboard aggregation; the original line is still stored as evidence.
    text = re.sub(
        r"[（(]\s*(?:"
        r"高性价比|高性价比之选|性价比|性价比首选|性价比之选|平价|平价多肽|高端|温和高端|"
        r"温和入门|入门|新手|日常|经典|修护|清爽|滋润|香氛|低敏|温和|均衡|全能|"
        r"家用|少刺激|干敏皮安心|敏感肌友好|敏感肌|干皮|油皮|口碑日牌|口碑爆款|"
        r"高销量|销量高|大容量热销|温和敏感眼|销量爆款|热销|爆款|首选|优选|推荐"
        r")\s*[^）)]*[）)]",
        "",
        text,
    )
    # Remove standalone capacity/quantity parentheticals like "(5ml)" or "(2ml×2)"
    # so the same product is not split across multiple rows.
    text = re.sub(r"[（(]\s*\d+(?:\.\d+)?\s*(?:ml|mL|ML|g|G|支|瓶|盒|片)(?:\s*×\s*\d+)?\s*[）)]", "", text)
    # Drop parenthetical transliterations/annotations such as "卡维拉（Cavilla）".
    text = re.sub(r"[（(]\s*[A-Za-z]+\s*[）)]", "", text)
    # Drop trailing capacity/quantity tokens like " 3ml" or " 5ml×2".
    text = re.sub(r"\s+\d+(?:\.\d+)?\s*(?:ml|mL|ML|g|G|支|瓶|盒|片)(?:\s*×\s*\d+)?$", "", text)
    # Drop unclosed trailing parentheticals such as "卡维拉睫毛精华液（精准纤长".
    text = re.sub(r"[（(].*$", "", text)
    text = re.sub(r"\s+", " ", text).strip(" -—：:，,。；;")
    if not any(k in text for k in PRODUCT_NAME_KEYWORDS):
        return ""
    # 卖点/功效通常是顿号列表，不是产品名；品牌+品名一般不会出现三个以上顿号。
    if text.count("、") >= 2:
        return ""
    # Reject ingredient descriptions like "2% 米诺地尔..." or bare
    # "米诺地尔 2%" that are not standalone product recommendations.
    if re.match(r"^\d+(?:\.\d+)?%\s*", text):
        return ""
    if re.match(r"^米诺地尔(?:\s*\d+%)?$", text):
        return ""
    if len(text) < 2 or len(text) > 36:
        return ""
    return text


def infer_brand_from_product_name(product_name):
    """Infer a stable master brand from a rule-extracted product name.

    When AI extraction is unavailable, the dashboard still needs brand-level
    aggregation. The master brand is normally the first token of the product
    name; a small list of known multi-token brand prefixes is handled
    explicitly so names like "梵玢 FBCY 睫毛臻萃精华液" keep both brand
    tokens while "朵妆 多肽睫毛滋养液" stays as brand=朵妆.
    """
    text = str(product_name or "").strip()
    if not text:
        return ""
    # Known multi-token brand prefixes that should stay together.
    KNOWN_BRAND_PREFIXES = (
        "梵玢 FBCY", "FBCY 梵玢", "梵玢", "FBCY",
        "RevitaLash", "Revitalash", "REVITALASH",
    )

    def _normalize_brand(brand):
        if brand.upper() == "REVITALASH":
            return "RevitaLash"
        if brand.upper() == "CAVILLA":
            return "卡维拉"
        if brand.lower() == "gerax":
            return "GeraX"
        if brand.upper() == "MAVALA":
            return "MAVALA"
        return brand

    lower_text = text.lower()
    for prefix in KNOWN_BRAND_PREFIXES:
        if lower_text.startswith(prefix.lower()):
            remainder = text[len(prefix):].strip()
            if remainder and any(k in remainder for k in PRODUCT_NAME_KEYWORDS):
                return _normalize_brand(prefix)
    # Match compact names against the shared known-brand list first, so
    # "施华蔻怡然染发霜" returns "施华蔻" instead of "施华蔻怡然".
    for brand in sorted(KNOWN_BRAND_NAMES, key=len, reverse=True):
        if lower_text.startswith(brand.lower()):
            remainder = text[len(brand):].strip()
            if remainder and any(k in remainder for k in PRODUCT_NAME_KEYWORDS):
                return _normalize_brand(brand)
    # For compact Chinese names like "卡维拉睫毛精华液" (no spaces), split at
    # the first product category keyword to obtain the brand.
    if " " not in text:
        first_kw_pos = min(
            (text.find(k) for k in PRODUCT_NAME_KEYWORDS if k in text),
            default=-1,
        )
        if first_kw_pos > 0:
            brand = text[:first_kw_pos].strip(" -—：:，,。；;")
            if brand and not brand.isdigit():
                text = brand
    # Default: first whitespace-separated token is the brand.
    tokens = text.split()
    brand = tokens[0] if tokens else text
    brand = re.sub(r"^[^\w\u4e00-\u9fff]+", "", brand)
    brand = re.sub(r"[^\w\u4e00-\u9fff]+$", "", brand)
    if not brand or brand.isdigit():
        return ""
    return _normalize_brand(brand)


def extract_products(answer_text):
    seen = set()
    products = []
    for line in split_answer_product_lines(answer_text):
        if any(hint in line for hint in ("#", "参考资料", "相关视频", "测评", "指南", "红黑榜", "一句话点评")):
            continue
        stripped = _line_after_leading_markers(line)
        # Skip leading explanation/warning lines that happen to contain product keywords.
        if re.match(r"^(?:选|挑选)\s*(?:睫毛增长液|眉毛增长液|染发剂|沐浴露|洗发水|面膜|面霜|防晒霜)", stripped):
            continue
        if re.match(r"^(?:先讲重点|先提醒|先划重点|先避雷|先避坑|先分清|先说|注意|提示|临床验证|米诺地尔属|米诺地尔效果|米诺地尔激活|米诺地尔有)", stripped):
            continue
        if re.match(r"^\d+\s*(?:款|个|种|类|支|瓶)\s*.*(?:增长液|精华液|滋养液|染发剂|面膜|沐浴露)", stripped):
            continue
        # Skip category-overview lines such as "温和肽类、米诺地尔（强效）".
        if re.search(r"温和肽类|猛药米诺地尔|米诺地尔\s*[（(]", line):
            continue
        # Skip summary/warning lines that merely describe a category or side effect.
        if re.search(r"(?:眉毛增长液|睫毛增长液|眉毛滋养液|睫毛滋养液|眉毛精华液|睫毛精华液).*?(?:需坚持|才见效|有泛红|有刺激|有干燥|有风险|属药品|是药品)", line):
            continue
        # Skip ingredient/detail lines that mention a product keyword but are
        # not themselves a product recommendation.
        if re.match(r"^(?:核心|核心成分|主要成分|配方|功效|适合人群|适用人群|价格|容量|参考价|用法|注意)\s*[：:]", stripped):
            continue
        if "≠" in line:
            continue
        if not any(k in line for k in PRODUCT_NAME_KEYWORDS):
            continue
        is_heading = (
            re.match(r"^[^\n]{0,4}(?:\d+|[一二三四五六七八九十])[\.\、]", line)
            or any(hint in line[:28] for hint in PRODUCT_HEADING_HINTS)
            or re.match(r"^[✨🌿🛡🏆💡⚠✅⭐💧🔥💰]", line)
            or ("｜" in line or "|" in line)
            or (
                (":" in line or "：" in line)
                and len(stripped.split("：" if "：" in stripped else ":", 1)[0]) <= 20
            )
            or (
                len(line) <= 60
                and not re.search(r"[。；;，,：:]", line)
            )
        )
        if not is_heading:
            continue
        name = normalize_known_product_alias(clean_recommended_product_name(line))
        if not name or name in seen:
            continue
        seen.add(name)
        products.append({
            "product_name": name,
            "brand_name": infer_brand_from_product_name(name),
            "evidence": line[:240],
        })
    if not products:
        text = str(answer_text or "")
        hits = []
        for tokens, canonical in PRODUCT_ALIAS_RULES:
            lower_text = text.lower()
            if not all(token.lower() in lower_text for token in tokens):
                continue
            positions = [lower_text.find(token.lower()) for token in tokens]
            pos = min(p for p in positions if p >= 0)
            start = max(0, pos - 80)
            end = min(len(text), pos + 180)
            hits.append((pos, canonical, text[start:end].strip()))
        for _pos, canonical, evidence in sorted(hits, key=lambda item: item[0]):
            key = canonical.lower()
            if key in seen:
                continue
            seen.add(key)
            products.append({
                "product_name": canonical,
                "brand_name": infer_brand_from_product_name(canonical),
                "evidence": evidence[:240],
            })
    return products


_BRAND_REGISTRY = {"loaded": False, "brands": set(), "mtime": 0}


def _load_historical_brand_registry():
    """Load master brands already captured in previous extractions.

    Rebuilds the registry when the products CSV changes so newly captured
    brands become available for fast text matching without restarting.
    """
    try:
        import doubao_dashboard_server as dashboard
    except Exception:
        return set()
    try:
        product_mtime = os.path.getmtime(OUT_PRODUCTS_CSV)
    except Exception:
        product_mtime = 0
    try:
        settings_mtime = brand_settings.SETTINGS_PATH.stat().st_mtime_ns
    except Exception:
        settings_mtime = 0
    mtime = (product_mtime, settings_mtime)
    if _BRAND_REGISTRY["loaded"] and _BRAND_REGISTRY["mtime"] == mtime:
        return _BRAND_REGISTRY["brands"]
    brands = set(dashboard.KNOWN_BRANDS)
    brands.update(
        item["name"] for item in brand_settings.vocabulary()
        if str(item.get("name") or "").strip()
    )
    if os.path.exists(OUT_PRODUCTS_CSV):
        with open(OUT_PRODUCTS_CSV, "r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                brand = str(row.get("brand_name") or "").strip()
                if not brand or dashboard.is_invalid_brand_candidate(brand):
                    continue
                canonical = dashboard.canonical_brand_name(brand)
                if canonical:
                    brands.add(canonical)
    _BRAND_REGISTRY.update({"loaded": True, "brands": brands, "mtime": mtime})
    debug_log("historical brand registry loaded; brands=" + str(len(brands)))
    return brands


def match_historical_brands(answer_text):
    """Fast-match already-known brands in the answer body.

    Each line that contains a known brand alias is turned into a product
    candidate with rank = appearance order.  This avoids calling the model
    when the answer only repeats previously captured brands, and provides a
    stable rank when the model is used as fallback.
    """
    text = strip_reference_prefix(answer_text)
    if not text:
        return []
    try:
        import doubao_dashboard_server as dashboard
    except Exception:
        return []
    brands = sorted(_load_historical_brand_registry(), key=len, reverse=True)
    if not brands:
        return []
    lines = split_answer_product_lines(text)
    seen = set()
    hits = []
    for line in lines:
        line_lower = line.lower()
        for brand in brands:
            aliases = set(dashboard.aliases_for_brand(brand))
            aliases.update(brand_settings.aliases_for_brand(brand))
            aliases = sorted(
                (alias for alias in aliases if alias),
                key=lambda value: (-len(value), value.casefold()),
            )
            if not any(alias.casefold() in line_lower for alias in aliases):
                continue
            cleaned = clean_recommended_product_name(line)
            if not cleaned:
                # Fallback: build a candidate from alias + tail of the line.
                for alias in aliases:
                    pos = line_lower.find(alias.casefold())
                    if pos >= 0:
                        tail = line[pos:]
                        tail = re.split(r"[。；;，,：:]", tail)[0]
                        cleaned = clean_recommended_product_name(tail)
                        if cleaned:
                            break
            if not cleaned:
                continue
            cleaned = normalize_known_product_alias(cleaned)
            key = cleaned.casefold()
            if key in seen:
                continue
            seen.add(key)
            hits.append({
                "product_name": cleaned,
                "brand_name": brand,
                "evidence": line[:240],
                "rank": len(hits) + 1,
                "rank_type": "appearance_order",
            })
            break
    return hits


def capture_metadata(payload):
    def value(name):
        current = payload.get(name)
        return "" if current is None else current

    captured_at = (
        value("captured_at")
        or value("capturedAt")
        or value("extractedAt")
    )
    return {
        "account_uid": value("account_uid"),
        "account_uid_masked": value("account_uid_masked"),
        "account_nickname": value("account_nickname"),
        "web_account_uid": value("web_account_uid"),
        "source_device": value("source_device"),
        "mumu_instance": value("mumu_instance"),
        "mumu_serial": value("mumu_serial"),
        "question_sent_at": beijing_time_str(value("question_sent_at")),
        "answer_completed_at": beijing_time_str(value("answer_completed_at")),
        "captured_at": beijing_time_str(captured_at, fallback_now=True),
        "source_uploaded_at": beijing_time_str(value("source_uploaded_at")),
        "receiver_received_at": beijing_time_str(value("receiver_received_at")),
    }


def normalize_payload_times(payload):
    captured_at = beijing_time_str(
        payload.get("captured_at")
        or payload.get("capturedAt")
        or payload.get("extractedAt"),
        fallback_now=True,
    )
    payload["captured_at"] = captured_at
    payload["extractedAt"] = captured_at
    for name in (
        "question_sent_at",
        "answer_completed_at",
        "source_uploaded_at",
        "receiver_received_at",
    ):
        if payload.get(name):
            payload[name] = beijing_time_str(payload.get(name))
    return payload


def ensure_named_csv_schema(path, fields, backup_suffix):
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        existing_fields = reader.fieldnames or []
        if existing_fields == fields:
            return
        rows = list(reader)
    backup = path + backup_suffix
    if not os.path.exists(backup):
        shutil.copy2(path, backup)
    fd, temp_path = tempfile.mkstemp(
        prefix="doubao_schema_",
        suffix=".csv",
        dir=BASE_DIR,
    )
    os.close(fd)
    try:
        with open(temp_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=fields,
                extrasaction="ignore",
            )
            writer.writeheader()
            for row in rows:
                item = {field: row.get(field, "") for field in fields}
                if "run_no" in fields and not item.get("run_no"):
                    item["run_no"] = row.get("run_no") or ""
                if "chat_id" in fields and not item.get("chat_id"):
                    item["chat_id"] = chat_id_from_url(
                        item.get("page_url") or row.get("page_url") or ""
                    )
                writer.writerow(item)
        os.replace(temp_path, path)
    finally:
        if os.path.exists(temp_path):
            os.unlink(temp_path)


def ensure_csv_schema():
    ensure_named_csv_schema(
        OUT_CSV,
        FIELDS,
        ".before_account_fields.bak",
    )


def ensure_products_csv_schema():
    """Add review metadata without discarding historical product rows."""
    ensure_named_csv_schema(
        OUT_PRODUCTS_CSV,
        PRODUCT_FIELDS,
        ".before_account_fields.bak",
    )


def ensure_answers_csv_schema():
    ensure_named_csv_schema(
        OUT_ANSWERS_CSV,
        ANSWER_FIELDS,
        ".before_account_fields.bak",
    )


def append_answer_csv(payload, run_no, run_time, answer_text, review_status, model):
    """Persist the exact answer body so every model decision is auditable/retriable."""
    normalize_payload_question(payload)
    ensure_answers_csv_schema()
    text = str(answer_text or "")
    digest = hashlib.sha256(text.encode("utf-8")).hexdigest() if text else ""
    with product_data_write_lock():
        exists = os.path.exists(OUT_ANSWERS_CSV)
        if exists:
            with open(OUT_ANSWERS_CSV, "r", encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    if str(row.get("run_no") or "") == str(run_no) and row.get("answer_hash") == digest:
                        debug_log("answer append deduplicated; run_no=" + str(run_no))
                        return digest
        with open(OUT_ANSWERS_CSV, "a", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=ANSWER_FIELDS)
            if not exists:
                writer.writeheader()
            writer.writerow({
                "run_no": run_no,
                "run_time": run_time,
                "chat_id": chat_id_from_url(payload.get("url", "")),
                "chat_title": payload.get("chatTitle") or payload.get("title") or "",
                "question": payload.get("question") or "",
                "page_url": payload.get("url") or "",
                **capture_metadata(payload),
                "answer_text": text,
                "answer_hash": digest,
                "review_status": review_status,
                "model": model,
                "reviewed_at": now_str(),
                "extracted_at": payload.get("extractedAt") or "",
            })
    return digest


def replace_file_with_retry(temp_path, target_path, attempts=12, delay_seconds=0.25):
    """Replace a CSV on Windows even when an antivirus/viewer holds it briefly.

    Returns False instead of raising after the bounded retry window.  A failed
    status write is safe because the answer body has already been archived as
    ai_pending and will be picked up by the background review worker.
    """
    last_error = None
    for attempt in range(attempts):
        try:
            os.replace(temp_path, target_path)
            return True
        except PermissionError as exc:
            last_error = exc
            if attempt + 1 < attempts:
                time.sleep(delay_seconds)
    debug_log("CSV replace deferred; target=" + target_path + " error=" + repr(last_error))
    return False


def update_answer_review(run_no, answer_hash, review_status, model):
    """Atomically update the pre-saved answer after a foreground/background review."""
    if not answer_hash or not os.path.exists(OUT_ANSWERS_CSV):
        return False
    try:
        with product_data_write_lock():
            with open(OUT_ANSWERS_CSV, "r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))
            changed = False
            for row in rows:
                if str(row.get("run_no") or "") == str(run_no) and row.get("answer_hash") == answer_hash:
                    row["review_status"] = review_status
                    row["model"] = model
                    row["reviewed_at"] = now_str()
                    changed = True
                    break
            if not changed:
                return False
            fd, temp_path = tempfile.mkstemp(prefix="doubao_answers_", suffix=".csv", dir=BASE_DIR)
            os.close(fd)
            try:
                with open(temp_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=ANSWER_FIELDS, extrasaction="ignore")
                    writer.writeheader()
                    writer.writerows(rows)
                if not replace_file_with_retry(temp_path, OUT_ANSWERS_CSV):
                    return False
                return True
            finally:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
    except TimeoutError as exc:
        debug_log("answer review status deferred: " + repr(exc))
        return False


def find_existing_source_run(payload):
    """Find a partially/successfully saved capture so retries are idempotent."""
    chat_id = chat_id_from_url(payload.get("url", ""))
    extracted_at = str(payload.get("extractedAt") or "").strip()
    if not chat_id or not extracted_at:
        return None
    # A zero-reference capture has no row in OUT_CSV, but its answer is already
    # persisted. Search both datasets so a later backend-reference recovery
    # repairs the original run instead of creating a duplicate dashboard run.
    for path in (OUT_CSV, OUT_ANSWERS_CSV):
        if not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    if (
                        str(row.get("chat_id") or "") == chat_id
                        and str(row.get("extracted_at") or "").strip() == extracted_at
                    ):
                        return {
                            "run_no": str(row.get("run_no") or ""),
                            "run_time": str(row.get("run_time") or "") or now_str(),
                        }
        except Exception:
            continue
    return None


def append_csv(payload):
    normalize_payload_times(payload)
    normalize_payload_question(payload)
    rows = payload.get("items") or []

    question = str(payload.get("question") or "").strip()
    if not question or question == "\u672a\u91c7\u96c6\u5230\u95ee\u9898":
        print("\n!!! \u8b66\u544a\uff1a\u95ee\u9898\u6587\u672c\u672a\u91c7\u96c6\u5230\uff0c\u8bf7\u68c0\u67e5 GET_QUESTION_JS \u662f\u5426\u751f\u6548 !!!")
        print("  chat_title: " + str(payload.get("chatTitle") or "")[:80])
        print("  page_url:   " + str(payload.get("url") or "")[:100])
        print("  question \u5c06\u4f7f\u7528\u56de\u9000\u503c: " + repr(question or "\u7a7a"))
        print("")

    with product_data_write_lock():
        ensure_csv_schema()
        existing_run = find_existing_source_run(payload)
        if existing_run:
            run_no = existing_run["run_no"]
            run_time = existing_run["run_time"]
        else:
            run_no = next_run_no()
            run_time = now_str()

        source_rows = [{
            "run_no": run_no,
            "run_time": run_time,
            "chat_id": chat_id_from_url(payload.get("url", "")),
            "chat_title": payload.get("chatTitle") or payload.get("title") or "",
            "question": payload.get("question") or "",
            "page_url": payload.get("url") or "",
            **capture_metadata(payload),
            "status": payload.get("status") or "",
            "complete": payload.get("complete"),
            "count": payload.get("count"),
            "expected_count": payload.get("expectedCount"),
            "index": item.get("index"),
            "title": item.get("title"),
            "href": item.get("href"),
            "source": item.get("source"),
            "extracted_at": payload.get("extractedAt") or "",
        } for item in rows]

        if existing_run:
            with open(OUT_CSV, "r", encoding="utf-8-sig", newline="") as f:
                current_rows = list(csv.DictReader(f))
            existing_source_rows = [
                row for row in current_rows
                if str(row.get("run_no") or "") == str(run_no)
            ]
            if len(source_rows) > len(existing_source_rows):
                merged_rows = [
                    row for row in current_rows
                    if str(row.get("run_no") or "") != str(run_no)
                ] + source_rows
                fd, temp_path = tempfile.mkstemp(
                    prefix="doubao_refs_recovery_",
                    suffix=".csv",
                    dir=BASE_DIR,
                )
                os.close(fd)
                try:
                    with open(temp_path, "w", encoding="utf-8-sig", newline="") as f:
                        writer = csv.DictWriter(f, fieldnames=FIELDS, extrasaction="ignore")
                        writer.writeheader()
                        writer.writerows(merged_rows)
                    os.replace(temp_path, OUT_CSV)
                finally:
                    if os.path.exists(temp_path):
                        os.unlink(temp_path)
                debug_log(
                    "source run repaired; run_no=%s old=%s new=%s"
                    % (run_no, len(existing_source_rows), len(source_rows))
                )
            else:
                debug_log("source append deduplicated; run_no=" + str(run_no))
        else:
            exists = os.path.exists(OUT_CSV)
            with open(OUT_CSV, "a", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=FIELDS)
                if not exists:
                    writer.writeheader()
                writer.writerows(source_rows)

    append_products_csv(payload, run_no, run_time)
    return run_no, len(rows)


def append_products_csv(payload, run_no, run_time):
    normalize_payload_question(payload)
    question = str(payload.get("question") or "").strip()
    answer_text = payload.get("answerText") or payload.get("answer_text") or ""
    # Persist first. If a model timeout or process interruption happens below,
    # the background retry worker still has the exact answer body to review.
    answer_hash = append_answer_csv(payload, run_no, run_time, answer_text, "ai_pending", product_ai_model_label())
    products, review_status, extraction_method, model = review_products_with_ai(answer_text)
    if not update_answer_review(run_no, answer_hash, review_status, model):
        debug_log(
            "answer review status deferred; run_no=" + str(run_no)
            + " status=" + str(review_status)
            + " (background worker will retry)"
        )
    if not is_recommendation_question(question):
        debug_log("product statistics skipped; non recommendation question=" + repr(question) + " run_no=" + str(run_no))
        return 0
    if not products:
        debug_log(
            "product append skipped; no products "
            + "run_no=" + str(run_no)
            + " question=" + repr(question)
            + " answer_len=" + str(len(str(answer_text or "")))
            + " ai_product=" + repr(os.environ.get("DOUBAO_USE_AI_PRODUCT", ""))
            + " ai_source=" + repr(os.environ.get("DOUBAO_USE_AI_SOURCE", ""))
            + " has_anthropic=" + str(bool(os.environ.get("ANTHROPIC_API_KEY")))
            + " has_openai=" + str(bool(os.environ.get("OPENAI_API_KEY")))
        )
        return 0

    reviewed_at = now_str()
    replacement_rows = []
    for index, item in enumerate(products, 1):
        product_name = infer_product_name_from_payload_context(
            item.get("product_name") or "",
            item.get("evidence") or "",
            payload,
        )
        replacement_rows.append({
            "run_no": run_no,
            "run_time": run_time,
            "chat_id": chat_id_from_url(payload.get("url", "")),
            "chat_title": payload.get("chatTitle") or payload.get("title") or "",
            "question": payload.get("question") or "",
            "page_url": payload.get("url") or "",
            **capture_metadata(payload),
            "product_index": item.get("rank") or index,
            "product_name": product_name,
            "brand_name": item.get("brand_name") or "",
            "evidence": item.get("evidence") or "",
            "product_count": len(products),
            "rank_type": item.get("rank_type") or "appearance_order",
            "extraction_method": extraction_method,
            "review_status": review_status,
            "model": model,
            "reviewed_at": reviewed_at,
            "answer_hash": answer_hash,
            "extracted_at": payload.get("extractedAt") or "",
        })

    # A duplicated foreground request or a foreground/background race must
    # replace the run snapshot, never append a second copy of it.
    with product_data_write_lock():
        ensure_products_csv_schema()
        existing_rows = []
        if os.path.exists(OUT_PRODUCTS_CSV):
            with open(OUT_PRODUCTS_CSV, "r", encoding="utf-8-sig", newline="") as f:
                existing_rows = list(csv.DictReader(f))
        retained = [
            row for row in existing_rows
            if str(row.get("run_no") or "") != str(run_no)
        ]
        fd, temp_path = tempfile.mkstemp(prefix="doubao_products_", suffix=".csv", dir=BASE_DIR)
        os.close(fd)
        try:
            with open(temp_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=PRODUCT_FIELDS, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(retained + replacement_rows)
            if not replace_file_with_retry(temp_path, OUT_PRODUCTS_CSV):
                return 0
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)
    return len(products)


def write_xlsx_from_csv():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception:
        return False

    if not os.path.exists(OUT_CSV):
        return False

    with open(OUT_CSV, "r", encoding="utf-8-sig", newline="") as f:
        data = list(csv.reader(f))

    if not data:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "refs"

    for row in data:
        ws.append(row)

    ws.freeze_panes = "A2"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 8,
        "B": 20,
        "C": 18,
        "D": 18,
        "E": 30,
        "F": 42,
        "G": 12,
        "H": 10,
        "I": 8,
        "J": 14,
        "K": 8,
        "L": 80,
        "M": 70,
        "N": 24,
        "O": 26,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    end_row = max(ws.max_row, 2)
    end_col = ws.max_column
    end_letter = get_column_letter(end_col)
    table = Table(displayName="DoubaoRefs", ref=f"A1:{end_letter}{end_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    add_source_sheets(wb, ws)

    try:
        wb.save(OUT_XLSX)
        return True
    except PermissionError:
        return False


def classify_source(href):
    parsed = urlparse(href or "")
    host = parsed.netloc.lower().split(":")[0]
    if not host:
        return OTHER, UNKNOWN, u(0x65e0, 0x6548, 0x94fe, 0x63a5), host

    meta = get_source_meta(href)
    ai_result = get_ai_source_classification(href, meta)
    if ai_result:
        return (
            ai_result.get("source_type") or ARTICLE,
            ai_result.get("media") or media_name_from_host(host),
            ai_result.get("note") or u(0x0041, 0x0049, 0x4fe1, 0x6e90, 0x5224, 0x65ad),
            host,
        )

    media = meta.get("site_name") or media_name_from_host(host)
    source_type = infer_source_type(href, meta)
    note = meta.get("note") or infer_note(source_type, meta)
    return source_type, media, note, host


def infer_source_type(href, meta):
    parsed = urlparse(href or "")
    host = parsed.netloc.lower()
    path = parsed.path.lower()
    content_type = (meta.get("content_type") or "").lower()
    page_type = (meta.get("og_type") or "").lower()
    title = (meta.get("title") or "").lower()

    if "pdf" in content_type or path.endswith(".pdf"):
        return ARTICLE
    if any(token in path for token in ("/video", "/v/", "/watch", "/shorts", "/share/video")):
        return VIDEO
    if page_type.startswith("video") or "video" in content_type:
        return VIDEO
    if any(token in host for token in ("taobao", "tmall", "jd.", "pinduoduo", "1688", "amazon")):
        return PRODUCT_PAGE
    if any(word in title for word in ("video", "shorts")):
        return VIDEO
    return ARTICLE


def infer_note(source_type, meta):
    if source_type == VIDEO:
        return VIDEO + u(0x94fe, 0x63a5)
    if source_type == PRODUCT_PAGE:
        return u(0x7535, 0x5546, 0x5546, 0x54c1, 0x9875)
    if meta.get("official"):
        return u(0x5b98, 0x65b9, 0x673a, 0x6784) + "/" + ARTICLE
    return ARTICLE


def get_ai_source_classification(href, meta):
    if os.environ.get("DOUBAO_USE_AI_SOURCE", "").strip() not in ("1", "true", "TRUE", "yes"):
        return None
    if not (os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("OPENAI_API_KEY")):
        return None

    cache = load_json_cache(SOURCE_AI_CACHE_JSON)
    host = urlparse(href or "").netloc.lower().split(":")[0]
    key = host or href or ""
    if key in cache:
        cached = cache[key]
        if os.environ.get("DOUBAO_REFRESH_WEAK_AI_SOURCE", "").strip() not in ("1", "true", "TRUE", "yes"):
            debug_log("AI cache hit: " + key)
            return cached
        if not is_weak_media_name(cached.get("media", ""), host):
            debug_log("AI cache hit: " + key)
            return cached
        debug_log("AI cache weak, refresh: " + key + " -> " + str(cached.get("media", "")))

    debug_log("AI classify start: " + key)
    result = call_anthropic_source_classifier(href, meta) or call_openai_source_classifier(href, meta)
    if result:
        cache[key] = result
        save_json_cache(SOURCE_AI_CACHE_JSON, cache)
        debug_log("AI classify done: " + key + " -> " + str(result.get("media", "")))
    else:
        fallback = fallback_ai_source_result(href, meta)
        if fallback:
            cache[key] = fallback
            save_json_cache(SOURCE_AI_CACHE_JSON, cache)
        debug_log("AI classify fallback: " + key)
        result = fallback
    return result


def fallback_ai_source_result(href, meta):
    host = urlparse(href or "").netloc.lower().split(":")[0]
    if not host:
        return None
    source_type = infer_source_type(href, meta)
    media = media_name_from_host(host)
    if not media or is_weak_media_name(media, host):
        if source_type == VIDEO:
            media = VIDEO
        elif source_type == PRODUCT_PAGE:
            media = PRODUCT_PAGE
        else:
            media = UNKNOWN + u(0x7f51, 0x7ad9)
    return {
        "source_type": source_type,
        "media": media[:80],
        "note": "AI timeout fallback; cached to avoid blocking main flow",
    }


def call_anthropic_source_classifier(href, meta):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return None

    # The project already uses a DeepSeek Anthropic-compatible endpoint for
    # brand review.  Keep the same default here so one ANTHROPIC_API_KEY works
    # consistently across the monitoring pipeline.
    base_url = os.environ.get("ANTHROPIC_BASE_URL", "https://api.deepseek.com/anthropic").rstrip("/")
    model = os.environ.get("MODEL_ID") or os.environ.get("ANTHROPIC_MODEL") or "deepseek-v4-pro"
    domain = urlparse(href or "").netloc
    prompt = build_source_prompt(href, meta)

    body = {
        "model": model,
        "max_tokens": 600,
        "temperature": 0,
        "messages": [
            {
                "role": "user",
                "content": json.dumps(prompt, ensure_ascii=False),
            }
        ],
    }

    try:
        req = urllib.request.Request(
            base_url + "/v1/messages",
            data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=env_int("DOUBAO_AI_TIMEOUT", 3)) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="ignore"))
        parts = data.get("content") or []
        content = ""
        for part in parts:
            if isinstance(part, dict) and part.get("type") == "text":
                content += part.get("text", "")
        parsed = parse_json_object(content)
    except Exception as exc:
        debug_log("Anthropic classify failed: " + domain + " | " + repr(exc))
        return None

    return normalize_ai_result(parsed, domain)


def call_openai_source_classifier(href, meta):
    api_key = os.environ.get("OPENAI_API_KEY")
    model = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
    if not api_key:
        return None

    prompt = build_source_prompt(href, meta)

    body = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": "You classify web reference sources. Output strict JSON only.",
            },
            {
                "role": "user",
                "content": json.dumps(prompt, ensure_ascii=False),
            },
        ],
        "temperature": 0,
    }

    try:
        req = urllib.request.Request(
            "https://api.openai.com/v1/chat/completions",
            data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
            headers={
                "Authorization": "Bearer " + api_key,
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=env_int("DOUBAO_AI_TIMEOUT", 3)) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="ignore"))
        content = data["choices"][0]["message"]["content"]
        parsed = parse_json_object(content)
    except Exception as exc:
        debug_log("OpenAI classify failed: " + urlparse(href or "").netloc + " | " + repr(exc))
        return None

    return normalize_ai_result(parsed, urlparse(href or "").netloc)


def build_source_prompt(href, meta):
    return {
        "task": "你是信源识别助手。请根据一个唯一域名及其样例页面信息，判断这个域名对应的信源类型和媒体/平台/官网名称。",
        "important": [
            "输入是按域名去重后的样例，不是让你分析单条内容观点。",
            "media 必须是人能看懂的信源名称，例如：抖音、什么值得买、咸宁日报、中国新闻网、百家号、今日头条某媒体、某某官网。",
            "不要只返回裸域名，除非无法从域名和页面信息推断名称。",
            "如果是新闻站/报纸/机构官网，请写出媒体或机构名称。",
            "如果是视频平台，source_type=视频，media=平台名。",
            "如果是电商商品页，source_type=商品页，media=电商平台或店铺/站点名。",
            "只返回 JSON，不要解释。",
        ],
        "allowed_source_type": [VIDEO, ARTICLE, PRODUCT_PAGE, OTHER],
        "input": {
            "sample_href": href,
            "domain": urlparse(href or "").netloc,
            "page_site_name": meta.get("site_name", ""),
            "page_title": meta.get("title", ""),
            "og_type": meta.get("og_type", ""),
            "content_type": meta.get("content_type", ""),
            "page_content_excerpt": meta.get("content_text", ""),
        },
        "output_schema": {
            "source_type": "视频|文章|商品页|其他",
            "media": "信源名称/媒体名称/平台名称/官网名称",
            "note": "简短中文判断依据",
        },
    }


def normalize_ai_result(parsed, domain):
    source_type = parsed.get("source_type")
    if source_type not in (VIDEO, ARTICLE, PRODUCT_PAGE, OTHER):
        source_type = ARTICLE
    media = str(parsed.get("media") or "").strip()
    note = str(parsed.get("note") or "").strip()
    if not media:
        return None
    if is_weak_media_name(media, domain):
        return None
    return {
        "source_type": source_type,
        "media": media[:80],
        "note": note[:120],
    }


def is_weak_media_name(media, domain=""):
    text = str(media or "").strip().lower()
    host = str(domain or "").strip().lower().split(":")[0]
    if not text:
        return True
    if "." in text:
        return True
    if host:
        labels = [part for part in host.split(".") if part and part not in ("www", "m", "wap", "pc", "post", "news", "baby")]
        if text == host or text in labels:
            return True
        if len(text) <= 10 and any(text == label or text in label for label in labels):
            return True
    ascii_like = re.fullmatch(r"[a-z0-9_-]{3,16}", text or "")
    if ascii_like:
        return True
    return False


def parse_json_object(text):
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.I).strip()
        text = re.sub(r"```$", "", text).strip()
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        text = match.group(0)
    return json.loads(text)


def parse_product_json(text):
    """Parse product-model JSON, accepting harmless trailing commas.

    This is deliberately narrow: malformed semantic content is still rejected
    and retried, so a broken model answer can never become product statistics.
    """
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.I).strip()
        text = re.sub(r"```$", "", text).strip()
    start, end = text.find("{"), text.rfind("}")
    if start >= 0 and end > start:
        text = text[start:end + 1]
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Common LLM formatting slip: a trailing comma before a closing list
        # or object. This transformation cannot invent any product content.
        repaired = re.sub(r",\s*([}\]])", r"\1", text)
        return json.loads(repaired)


def build_source_prompt(href, meta):
    examples = [
        DOUYIN,
        u(0x4ec0, 0x4e48, 0x503c, 0x5f97, 0x4e70),
        XN_DAILY,
        CHINA_NEWS,
        u(0x767e, 0x5bb6, 0x53f7),
        u(0x4eca, 0x65e5, 0x5934, 0x6761),
        u(0x5b98, 0x7f51),
    ]
    return {
        "task": "Classify one unique source domain for a Chinese spreadsheet.",
        "important": [
            "Input is one deduplicated domain plus one sample URL/page metadata.",
            "Return JSON only.",
            "media must be a human-readable Chinese source/platform/official organization name whenever inferable, not merely the raw domain.",
            "Never use raw domain labels or short technical names as media, for example hfg386, dianlinet, hzpwjc, ifeng, smzdm. Infer a readable Chinese media/platform/shop/organization name from title, site name, domain meaning, and page context.",
            "If the page is a small ecommerce shop with no public media brand, return the shop/site display name in Chinese when possible, otherwise return a readable label such as 独立商城 or 品牌商城, not the bare domain.",
            "Examples of media names: " + ", ".join(examples),
            "For news/article sites, identify the news organization, newspaper, Baijiahao/Toutiao account, official organization, or website name.",
            "For video platforms, source_type must be the Chinese value for video and media should be platform/source name.",
            "For ecommerce product pages, source_type must be the Chinese value for product page and media should be ecommerce platform/store/site name.",
        ],
        "allowed_source_type": [VIDEO, ARTICLE, PRODUCT_PAGE, OTHER],
        "input": {
            "sample_href": href,
            "domain": urlparse(href or "").netloc,
            "page_site_name": meta.get("site_name", ""),
            "page_title": meta.get("title", ""),
            "og_type": meta.get("og_type", ""),
            "content_type": meta.get("content_type", ""),
            "page_content_excerpt": meta.get("content_text", ""),
        },
        "output_schema": {
            "source_type": "|".join([VIDEO, ARTICLE, PRODUCT_PAGE, OTHER]),
            "media": "source/media/platform/official organization name",
            "note": "short Chinese reason",
        },
    }


def media_name_from_host(host):
    host = (host or "").lower().split(":")[0]
    if host.startswith("www."):
        host = host[4:]
    if not host:
        return UNKNOWN
    parts = host.split(".")
    if len(parts) >= 3 and parts[-2] in ("com", "net", "org", "gov", "edu"):
        root = parts[-3]
    elif len(parts) >= 2:
        root = parts[-2]
    else:
        root = parts[0]
    return root or host


def load_source_cache():
    return load_json_cache(SOURCE_CACHE_JSON)


def save_source_cache(cache):
    save_json_cache(SOURCE_CACHE_JSON, cache)


def load_json_cache(path):
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_json_cache(path, cache):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def get_source_content_text(href, max_chars=3000):
    """Read archived body text from the content worker database, if available."""
    if not href or not os.path.exists(SOURCE_CONTENT_DB):
        return ""
    try:
        with sqlite3.connect(SOURCE_CONTENT_DB, timeout=5) as conn:
            row = conn.execute(
                "SELECT content_text FROM source_content WHERE url=? AND status='ok'",
                (href,),
            ).fetchone()
            if row and row[0]:
                return str(row[0])[:max_chars].strip()
    except Exception as exc:
        debug_log("Content db read failed: " + href + " | " + repr(exc))
    return ""


def get_source_meta(href):
    cache = load_source_cache()
    if href in cache:
        return cache[href]

    parsed = urlparse(href or "")
    host = parsed.netloc.lower().split(":")[0]
    meta = {
        "site_name": "",
        "title": "",
        "og_type": "",
        "content_type": "",
        "official": is_official_url(href),
        "note": "",
        "content_text": "",
    }

    if not href:
        return meta

    # Prefer the body already archived by the content worker.
    meta["content_text"] = get_source_content_text(href)

    try:
        debug_log("Meta fetch start: " + href)
        req = urllib.request.Request(
            href,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            },
            method="GET",
        )
        with urllib.request.urlopen(req, timeout=env_int("DOUBAO_META_TIMEOUT", 2)) as resp:
            meta["content_type"] = resp.headers.get("Content-Type", "")
            raw = resp.read(250000)
        debug_log("Meta fetch done: " + href)
    except Exception as exc:
        debug_log("Meta fetch failed: " + href + " | " + repr(exc))
        meta["site_name"] = media_name_from_host(host)
        cache[href] = meta
        save_source_cache(cache)
        return meta

    text = decode_html(raw)
    meta["site_name"] = (
        extract_meta(text, "property", "og:site_name")
        or extract_meta(text, "name", "application-name")
        or extract_meta(text, "name", "SiteName")
        or ""
    ).strip()
    meta["og_type"] = (extract_meta(text, "property", "og:type") or "").strip()
    meta["title"] = (
        extract_meta(text, "property", "og:title")
        or extract_title(text)
        or ""
    ).strip()

    if not meta["site_name"]:
        meta["site_name"] = infer_site_name_from_title(meta["title"]) or media_name_from_host(host)

    cache[href] = meta
    save_source_cache(cache)
    return meta


def is_official_url(href):
    parsed = urlparse(href or "")
    text = "." + parsed.netloc.lower() + parsed.path.lower()
    return any(hint in text for hint in OFFICIAL_HINTS)


def decode_html(raw):
    for enc in ("utf-8", "gb18030", "gbk", "big5", "latin1"):
        try:
            return raw.decode(enc, errors="ignore")
        except Exception:
            pass
    return raw.decode("utf-8", errors="ignore")


def extract_meta(html, attr_name, attr_value):
    attr_re = re.escape(attr_name)
    value_re = re.escape(attr_value)
    patterns = [
        rf"<meta[^>]+{attr_re}=[\"']{value_re}[\"'][^>]+content=[\"']([^\"']+)[\"'][^>]*>",
        rf"<meta[^>]+content=[\"']([^\"']+)[\"'][^>]+{attr_re}=[\"']{value_re}[\"'][^>]*>",
    ]
    for pattern in patterns:
        match = re.search(pattern, html, flags=re.I | re.S)
        if match:
            return html_unescape(match.group(1))
    return ""


def extract_title(html):
    match = re.search(r"<title[^>]*>(.*?)</title>", html, flags=re.I | re.S)
    if not match:
        return ""
    return html_unescape(re.sub(r"\s+", " ", match.group(1)).strip())


def html_unescape(text):
    try:
        import html
        return html.unescape(text or "")
    except Exception:
        return text or ""


def infer_site_name_from_title(title):
    title = (title or "").strip()
    if not title:
        return ""
    for sep in (" - ", "_", "|", "—", "-"):
        parts = [p.strip() for p in title.split(sep) if p.strip()]
        if len(parts) >= 2 and len(parts[-1]) <= 30:
            return parts[-1]
    return ""


def add_source_sheets(wb, refs_ws):
    headers = [cell.value for cell in refs_ws[1]]
    col = {name: index + 1 for index, name in enumerate(headers)}

    for name in ["source_analysis", "source_summary"]:
        if name in wb.sheetnames:
            del wb[name]

    rows = []
    for r in range(2, refs_ws.max_row + 1):
        href = refs_ws.cell(r, col["href"]).value
        source_type, media, note, domain = classify_source(href)
        rows.append({
            "run_no": refs_ws.cell(r, col["run_no"]).value,
            "index": refs_ws.cell(r, col["index"]).value,
            "source_type": source_type,
            "media": media,
            "domain": domain,
            "title": refs_ws.cell(r, col["title"]).value,
            "href": href,
            "note": note,
            "chat_id": refs_ws.cell(r, col["chat_id"]).value,
            "page_url": refs_ws.cell(r, col["page_url"]).value,
        })

    analysis = wb.create_sheet("source_analysis")
    analysis_headers = ["run_no", "index", "source_type", "media", "domain", "title", "href", "note", "chat_id", "page_url"]
    analysis.append(analysis_headers)
    for item in rows:
        analysis.append([item[h] for h in analysis_headers])

    summary = wb.create_sheet("source_summary")
    summary.append(["summary_type", "source_type", "media", "domain", "total_refs", "unique_links"])

    by_type = {}
    by_media = {}
    for item in rows:
        by_type.setdefault(item["source_type"], []).append(item)
        by_media.setdefault((item["source_type"], item["media"], item["domain"]), []).append(item)

    for source_type, items in sorted(by_type.items()):
        summary.append([BY_TYPE, source_type, "", "", len(items), len(set(i["href"] for i in items))])

    for (source_type, media, domain), items in sorted(by_media.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        summary.append([BY_MEDIA, source_type, media, domain, len(items), len(set(i["href"] for i in items))])

    format_worksheet(analysis, {"A": 8, "B": 8, "C": 12, "D": 36, "E": 26, "F": 70, "G": 72, "H": 36, "I": 20, "J": 44})
    format_worksheet(summary, {"A": 14, "B": 12, "C": 44, "D": 28, "E": 12, "F": 12})
    add_excel_table(analysis, "SourceAnalysis")
    add_excel_table(summary, "SourceSummary")


def format_worksheet(ws, widths):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.freeze_panes = "A2"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_excel_table(ws, table_name):
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    end_row = max(ws.max_row, 2)
    end_col = ws.max_column
    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(end_col)}{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def main():
    if len(sys.argv) >= 2:
        raw = sys.argv[1]
    else:
        raw = sys.stdin.read()

    payload = load_payload(raw)
    run_no, rows_written = append_csv(payload)
    has_xlsx = False
    if os.environ.get("DOUBAO_SKIP_XLSX", "").strip() not in ("1", "true", "TRUE", "yes"):
        has_xlsx = write_xlsx_from_csv()

    print(json.dumps({
        "ok": True,
        "run_no": run_no,
        "rows_written": rows_written,
        "csv": OUT_CSV,
        "xlsx": OUT_XLSX if has_xlsx else "",
        "count": payload.get("count", 0),
        "complete": payload.get("complete", False),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
