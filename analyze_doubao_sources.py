from collections import defaultdict
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import save_doubao_refs


PATH = r"C:\Users\AMD\Desktop\monitor\doubao_refs_result.xlsx"
OUT_PATH = r"C:\Users\AMD\Desktop\monitor\doubao_refs_result_source_analysis.xlsx"


def u(*codes):
    return "".join(chr(code) for code in codes)


VIDEO = u(0x89c6, 0x9891)
ARTICLE = u(0x6587, 0x7ae0)
PRODUCT_PAGE = u(0x5546, 0x54c1, 0x9875)
OTHER = u(0x5176, 0x4ed6)
UNKNOWN = u(0x672a, 0x77e5)
DOUYIN = u(0x6296, 0x97f3)
DOUYIN_VIDEO = DOUYIN + VIDEO + u(0x94fe, 0x63a5)
MEDIA = u(0x5a92, 0x4f53)
NEWS_MEDIA = u(0x65b0, 0x95fb) + MEDIA
CHINA_NEWS = u(0x4e2d, 0x56fd, 0x65b0, 0x95fb, 0x7f51)
XN_DAILY = u(0x54b8, 0x5b81, 0x65e5, 0x62a5)
HEALTH_CHINA = u(0x5065, 0x5eb7, 0x4e2d, 0x56fd, 0x7f51)
TAOBAO = u(0x6dd8, 0x5b9d)
BY_TYPE = u(0x6309, 0x7c7b, 0x578b, 0x6c47, 0x603b)
BY_MEDIA = u(0x6309, 0x5a92, 0x4f53, 0x6c47, 0x603b)


SOURCE_MAP = {
    "www.iesdouyin.com": (VIDEO, DOUYIN, DOUYIN_VIDEO),
    "iesdouyin.com": (VIDEO, DOUYIN, DOUYIN_VIDEO),
    "www.goodhousekeeping.com": (ARTICLE, "Good Housekeeping", u(0x6d77, 0x5916, 0x751f, 0x6d3b, 0x65b9, 0x5f0f) + "/" + u(0x4ea7, 0x54c1, 0x8bc4, 0x6d4b) + MEDIA),
    "goodhousekeeping.com": (ARTICLE, "Good Housekeeping", u(0x6d77, 0x5916, 0x751f, 0x6d3b, 0x65b9, 0x5f0f) + "/" + u(0x4ea7, 0x54c1, 0x8bc4, 0x6d4b) + MEDIA),
    "www.chinanews.com": (ARTICLE, CHINA_NEWS, NEWS_MEDIA),
    "chinanews.com": (ARTICLE, CHINA_NEWS, NEWS_MEDIA),
    "vigilanses.anses.fr": (ARTICLE, "Vigil'Anses / ANSES", u(0x6cd5, 0x56fd, 0x5b98, 0x65b9, 0x673a, 0x6784, 0x98ce, 0x9669, 0x8b66, 0x6212) + "/PDF"),
    "szb.xnnews.com.cn": (ARTICLE, XN_DAILY, u(0x5730, 0x65b9) + NEWS_MEDIA + "/" + u(0x6570, 0x5b57, 0x62a5)),
    "mtest.health-china.com": (ARTICLE, HEALTH_CHINA, u(0x5065, 0x5eb7, 0x8d44, 0x8baf) + MEDIA),
    "pcdetail.taobao.com": (PRODUCT_PAGE, TAOBAO, u(0x7535, 0x5546, 0x5546, 0x54c1, 0x9875) + "," + u(0x975e) + MEDIA + ARTICLE),
}


def classify(href):
    return save_doubao_refs.classify_source(href)


def style_sheet(ws):
    ws.freeze_panes = "A2"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_table(ws, table_name):
    end_row = max(ws.max_row, 2)
    end_col = ws.max_column
    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(end_col)}{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def main():
    wb = load_workbook(PATH)
    refs = wb["refs"]
    headers = [c.value for c in refs[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    rows = []
    for r in range(2, refs.max_row + 1):
        href = refs.cell(r, idx["href"]).value
        source_type, media, note, domain = classify(href)
        rows.append({
            "run_no": refs.cell(r, idx["run_no"]).value,
            "index": refs.cell(r, idx["index"]).value,
            "source_type": source_type,
            "media": media,
            "domain": domain,
            "title": refs.cell(r, idx["title"]).value,
            "href": href,
            "note": note,
            "chat_id": refs.cell(r, idx["chat_id"]).value,
            "page_url": refs.cell(r, idx["page_url"]).value,
        })

    for name in ["source_analysis", "source_summary"]:
        if name in wb.sheetnames:
            del wb[name]

    analysis = wb.create_sheet("source_analysis")
    analysis_headers = ["run_no", "index", "source_type", "media", "domain", "title", "href", "note", "chat_id", "page_url"]
    analysis.append(analysis_headers)
    for item in rows:
        analysis.append([item[h] for h in analysis_headers])

    summary = wb.create_sheet("source_summary")
    summary.append(["summary_type", "source_type", "media", "domain", "total_refs", "unique_links"])

    by_type = defaultdict(list)
    by_media = defaultdict(list)
    for item in rows:
        by_type[item["source_type"]].append(item)
        by_media[(item["source_type"], item["media"], item["domain"])].append(item)

    for source_type, items in sorted(by_type.items()):
        summary.append([BY_TYPE, source_type, "", "", len(items), len(set(i["href"] for i in items))])

    for (source_type, media, domain), items in sorted(by_media.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        summary.append([BY_MEDIA, source_type, media, domain, len(items), len(set(i["href"] for i in items))])

    style_sheet(analysis)
    style_sheet(summary)

    for col, width in {"A": 8, "B": 8, "C": 12, "D": 36, "E": 26, "F": 70, "G": 72, "H": 36, "I": 20, "J": 44}.items():
        analysis.column_dimensions[col].width = width
    for col, width in {"A": 14, "B": 12, "C": 44, "D": 28, "E": 12, "F": 12}.items():
        summary.column_dimensions[col].width = width

    add_table(analysis, "SourceAnalysis")
    add_table(summary, "SourceSummary")

    try:
        wb.save(PATH)
    except PermissionError:
        wb.save(OUT_PATH)


if __name__ == "__main__":
    main()
