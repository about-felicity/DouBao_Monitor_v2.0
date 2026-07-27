"""
监控表数据处理脚本
输入：监控表.xlsx
输出：监控结果.xlsx
"""

import re
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
import os
import sys


def parse_product_info(text: str) -> dict:
    """从 B 列原始文本中提取各字段"""
    first_line = text.split("\n")[0].strip()

    # 日期：MM/DD 格式
    date_match = re.search(r"\b(\d{2}/\d{2})\b", first_line)
    date = date_match.group(1) if date_match else ""

    # 抖音链接
    url_match = re.search(r"(https://v\.douyin\.com/[^\s/]+/?)", first_line)
    url = url_match.group(1) if url_match else ""

    # 产品描述：链接之后的文字
    if url:
        after_url = first_line[first_line.index(url) + len(url):].strip()
        # 去掉【抖音商城】前缀（有时出现在链接之前），取链接后的描述
        description = after_url
    else:
        description = ""

    return {
        "日期": date,
        "抖音链接": url,
        "产品描述": description,
    }


def process_monitoring_excel(input_file: str, output_file: str) -> pd.DataFrame:
    """读取监控表，结构化后写出到新 Excel"""
    df_raw = pd.read_excel(input_file, sheet_name=0, header=None)

    records = []
    current_store = ""

    for _, row in df_raw.iterrows():
        col_a = row[0] if pd.notna(row[0]) else None
        col_b = row[1] if pd.notna(row[1]) else None

        # A 列有值 → 更新当前店铺名
        if col_a:
            current_store = str(col_a).strip()

        # B 列有值 → 这是一条产品记录
        if col_b:
            info = parse_product_info(str(col_b))
            info["店铺名称"] = current_store
            records.append(info)

    result_df = pd.DataFrame(
        records,
        columns=["店铺名称", "日期", "抖音链接", "产品描述"],
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="监控结果")

        ws = writer.sheets["监控结果"]

        # 列宽
        col_widths = {
            "A": 22,  # 店铺名称
            "B": 8,   # 日期
            "C": 45,  # 抖音链接
            "D": 50,  # 产品描述
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        # 数据行：自动换行关闭，垂直居中
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)

        # 冻结首行
        ws.freeze_panes = "A2"

    print(f"处理完成！共 {len(records)} 条记录 → {output_file}")
    return result_df


if __name__ == "__main__":
    input_file = os.path.join(os.path.dirname(__file__), "监控表.xlsx")
    output_file = os.path.join(os.path.dirname(__file__), "监控结果.xlsx")

    if not os.path.exists(input_file):
        print(f"找不到输入文件：{input_file}")
        sys.exit(1)

    result = process_monitoring_excel(input_file, output_file)
    print("\n预览：")
    print(result.to_string(index=False))
